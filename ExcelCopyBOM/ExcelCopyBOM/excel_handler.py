"""
Excel handling for ExcelCopyBOM using pandas and openpyxl
"""
import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
from pathlib import Path
import re
import uuid
import shutil
import logging
import win32com.client
import pythoncom
from typing import Dict, List, Optional, Tuple
from . import config
from .database import DrawingDatabase
from image_handler import ExcelImageHandler
from PIL import ImageGrab
import tempfile
import win32clipboard

class ExcelHandler:
    def __init__(self, bom_path: str, prev_bom_path: Optional[str] = None):
        self.bom_path = Path(bom_path)
        self.prev_bom_path = Path(prev_bom_path) if prev_bom_path else None
        self.db = DrawingDatabase()
        
        # Excel workbooks
        self.wb = None  # openpyxl workbook
        self.df_bom = None  # pandas dataframe for BOM
        self.excel = None  # COM object
        
        # Extracted metadata
        self.part_number = None
        self.revision = None
        self.required_columns = {}  # column name -> index mapping
        
        # Billede håndtering
        self.image_mapping = {}  # part_number -> temp_image_path
        self.temp_images = []  # Liste over midlertidige billeder der skal slettes
        self.image_handler = ExcelImageHandler()
        
    def _init_excel_com(self):
        """Initialiser Excel COM objekt"""
        if self.excel is None:
            pythoncom.CoInitialize()
            self.excel = win32com.client.Dispatch("Excel.Application")
            self.excel.Visible = False
            self.excel.DisplayAlerts = False
            
    def _cleanup_excel_com(self):
        """Luk Excel COM objekt"""
        if self.excel:
            try:
                self.excel.Quit()
            except:
                pass
            finally:
                self.excel = None
                pythoncom.CoUninitialize()
                
    def _insert_arrangement_row(self):
        """Indsæt arrangement række i row 2"""
        arrangement_data = {
            'Item': 1,
            'Qty': 1,
            'Part Number': self.part_number,
            'Description': f"Arrangement Drawing {self.part_number}",
            'REV': self.revision,
            'Category': 'Area Arrangement Drawing'
        }
        
        # Indsæt ny række i pandas DataFrame
        self.df_bom.loc[1.5] = arrangement_data  # Brug 1.5 for at sikre rigtig position efter sortering
        self.df_bom = self.df_bom.sort_index().reset_index(drop=True)
        
    def _delete_supplier_parts(self):
        """Slet leverandør dele"""
        supplier_patterns = ['0000-700', '0000-701', '0000-702']
        mask = ~self.df_bom['Part Number'].str.contains('|'.join(supplier_patterns), na=False)
        self.df_bom = self.df_bom[mask].reset_index(drop=True)
        
    def _update_revisions(self):
        """Opdater REV kolonne med seneste tegningsrevisioner"""
        for idx, row in self.df_bom.iterrows():
            part_num = str(row['Part Number'])
            if pd.isna(part_num) or not part_num.strip():
                continue
                
            # Find seneste revision fra databasen
            latest_rev = self.db.find_latest_revision(part_num)
            if latest_rev:
                self.df_bom.at[idx, 'REV'] = latest_rev
                
            # Tilføj tegningsstatus
            drawing_status = self.db.get_drawing_status(part_num)
            self.df_bom.at[idx, 'Drawing Status'] = drawing_status
            
    def _extract_images(self):
        """Udtræk billeder fra Thumbnail kolonnen"""
        logging.info("Starter udtrækning af billeder")
        
        try:
            # Initialiser Excel COM
            self._init_excel_com()
            
            # Åbn workbook
            wb = self.excel.Workbooks.Open(str(self.bom_path.absolute()))
            sheet = wb.Sheets(1)
            
            # Find Thumbnail og Part Number kolonner
            thumbnail_col = None
            part_num_col = None
            for col in range(1, sheet.UsedRange.Columns.Count + 1):
                header = sheet.Cells(1, col).Value
                if header == "Thumbnail":
                    thumbnail_col = col
                elif header == "Part Number":
                    part_num_col = col
                    
            if not thumbnail_col or not part_num_col:
                logging.warning("Kunne ikke finde nødvendige kolonner")
                return
                
            logging.debug(f"Thumbnail kolonne: {thumbnail_col}, Part Number kolonne: {part_num_col}")
            
            # Gennemgå alle shapes
            for shape in sheet.Shapes:
                if shape.Type == 13:  # 13 = billede
                    # Find cellen som billedet er i
                    cell_row = int((shape.TopLeftCell.Row + shape.BottomRightCell.Row) / 2)
                    cell_col = shape.TopLeftCell.Column
                    
                    # Hvis billedet er i Thumbnail kolonnen
                    if cell_col == thumbnail_col and cell_row > 1:  # Skip header
                        # Få part number fra samme række
                        part_num = str(sheet.Cells(cell_row, part_num_col).Value)
                        if not part_num or part_num == "None":
                            continue
                            
                        logging.debug(f"Fandt billede for {part_num} i række {cell_row}")
                        
                        # Gem billede til temp fil
                        temp_path = config.TEMP_DIR / f"{uuid.uuid4()}.png"
                        shape.Copy()  # Kopier til clipboard
                        
                        # Gem fra clipboard som billede
                        try:
                            img = ImageGrab.grabclipboard()
                            if img:
                                img.save(temp_path)
                                self.image_mapping[part_num] = temp_path
                                self.temp_images.append(temp_path)
                                logging.info(f"Gemt billede for {part_num} til {temp_path}")
                        except Exception as e:
                            logging.warning(f"Kunne ikke gemme billede for {part_num}: {e}")
                            
            # Luk workbook uden at gemme
            wb.Close(False)
            
        except Exception as e:
            logging.exception(f"Fejl under udtrækning af billeder: {e}")
            
        finally:
            self._cleanup_excel_com()
            
    def _handle_images(self, output_path: str):
        """Håndter billeder ved hjælp af ExcelImageHandler"""
        if not self.image_mapping:
            logging.info("Ingen billeder at håndtere")
            return
            
        logging.info(f"Gemmer workbook uden billeder til {output_path}")
        # Først gem workbook uden billeder
        self.wb.save(output_path)
        
        logging.info("Tilføjer billeder via COM")
        # Derefter tilføj billeder via COM
        success = self.image_handler.process_images(
            Path(output_path),
            self.image_mapping
        )
        
        if not success:
            logging.warning("Kunne ikke indsætte alle billeder korrekt")
        else:
            logging.info("Billeder indsat succesfuldt")
            
    def load_workbook(self):
        """Load Excel workbook and extract metadata"""
        logging.info(f"Indlæser Excel fil: {self.bom_path}")
        
        # Load with pandas for data manipulation
        self.df_bom = pd.read_excel(self.bom_path, engine='openpyxl')
        logging.debug(f"DataFrame kolonner: {self.df_bom.columns.tolist()}")
        
        # Load with openpyxl for formatting
        self.wb = openpyxl.load_workbook(self.bom_path)
        logging.debug(f"Worksheet navne: {self.wb.sheetnames}")
        
        # Extract part number and revision from filename
        filename = self.bom_path.stem
        match = re.match(r"(\d{4}-\d{2}\.\d(?:-[A-Z]\d{2})?)-([A-Z]\d{2})?", filename)
        if match:
            self.part_number = match.group(1)
            self.revision = match.group(2) or ""
            logging.debug(f"Extracted: Part Number={self.part_number}, REV={self.revision}")
            
    def process_bom(self):
        """Hovedfunktion for BOM behandling"""
        # Udtræk billeder før andre ændringer
        self._extract_images()
        
        # Eksisterende processer
        self._insert_arrangement_row()
        self._delete_supplier_parts()
        self._update_revisions()
        
    def save_workbook(self, output_path: str):
        """Gem Excel fil med alle ændringer"""
        try:
            # Håndter billeder via COM
            self._handle_images(output_path)
        finally:
            # Ryd altid op i temp filer
            self._cleanup_temp_files()
            
    def _cleanup_temp_files(self):
        """Ryd op i midlertidige billeder"""
        logging.info(f"Rydder op i {len(self.temp_images)} midlertidige filer")
        for temp_file in self.temp_images:
            try:
                temp_file.unlink()
                logging.debug(f"Slettet: {temp_file}")
            except Exception as e:
                logging.warning(f"Kunne ikke slette temp fil {temp_file}: {e}")
                
    def __del__(self):
        """Destructor der sikrer oprydning"""
        self._cleanup_temp_files()
        self._cleanup_excel_com() 