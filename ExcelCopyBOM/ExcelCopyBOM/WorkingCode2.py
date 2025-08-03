import os
import shutil
import pandas as pd
import openpyxl
import time
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font
from tkinter import Tk, filedialog, messagebox
from win32com.client import Dispatch
from concurrent.futures import ThreadPoolExecutor
from Categories import PartNumberParser
from tkinter import ttk
import tkinter as tk
import excel_formatting
import excel_grouping
from openpyxl import load_workbook
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
from openpyxl.drawing.image import Image
from openpyxl.utils.units import pixels_to_EMU, cm_to_EMU
import pythoncom

def choose_file():
    """Lader brugeren vælge en Excel-fil og returnerer dens sti og part number."""
    root = Tk()
    root.withdraw()
    # file_path = filedialog.askopenfilename(title="Open Excel BOM file",
    #                                        filetypes=[("Excel files", ".xlsx .xls")],
    #                                        initialdir="C:\\Working Folder\\Designs\\5-Projects")
    file_path = "C:\\Coding\\Python\\ExcelCopyBOM\\4003-02.1-A01-- - BOM.xlsx"
    if not file_path:
        os._exit(1)
    
    # Split filnavnet for at få part number
    file_name = os.path.basename(file_path)
    part_number = file_name.split(" - ")[0]
    
    # Brug Categories.py til at finde kategorien
    parser = PartNumberParser()
    category = parser.find_category(part_number)
    
    print(f"Part Number: {part_number}")
    print(f"Category: {category}")
    
    return file_path, part_number, category

def create_copy(file_path):
    """Opretter en kopi af den originale Excel-fil i Files mappen."""
    # Opret Files mappe hvis den ikke findes
    files_dir = os.path.join(os.path.dirname(file_path), "Files")
    os.makedirs(files_dir, exist_ok=True)
    
    # Opret kopi i Files mappen
    file_name = os.path.basename(file_path)
    dest_path = os.path.join(files_dir, file_name)
    shutil.copy2(file_path, dest_path)
    return dest_path

def load_categories():
    """Indlæser kategorier fra en ekstern .txt-fil."""
    categories = {}
    categories_file = "categories.txt"
    if os.path.exists(categories_file):
        with open(categories_file, "r") as file:
            for line in file:
                parts = line.strip().split("=")
                if len(parts) == 2:
                    categories[parts[0].strip()] = parts[1].strip()
    return categories

def categorize_data(df):
    """Opdeler data i kategorier baseret på Part Number (kolonne 2) og returnerer en dict med kategorier."""
    parser = PartNumberParser()
    categorized_data = {}
    
    for _, row in df.iterrows():
        part_number = str(row.iloc[1])  # Kolonne 2 (Part Number)
        category = parser.find_category(part_number)
        
        if category not in categorized_data:
            categorized_data[category] = pd.DataFrame(columns=df.columns)
        categorized_data[category] = pd.concat([categorized_data[category], row.to_frame().T], ignore_index=True)
    
    return categorized_data

def write_debug(message):
    """Skriver debug beskeder til Debug.txt"""
    debug_file = os.path.join(os.path.dirname(__file__), "Debug.txt")
    timestamp = time.strftime("%Y-%m-%d %H:%M:%S")
    with open(debug_file, 'a', encoding='utf-8') as f:
        f.write(f"{timestamp}: {message}\n")

def ensure_file_access(file_path):
    """Sikrer at filen ikke er låst af Excel og har de rette tilladelser."""
    try:
        # Tjek om filen eksisterer
        if not os.path.exists(file_path):
            return True
            
        # Prøv at åbne filen for at tjekke tilladelser
        with open(file_path, 'a+b') as f:
            f.seek(0)
        return True
    except PermissionError:
        write_debug(f"Permission denied for file: {file_path}")
        # Prøv at lukke alle Excel-processer
        try:
            os.system("taskkill /f /im excel.exe")
            time.sleep(2)  # Vent på at processerne lukkes
            write_debug("Forced close of Excel processes")
            return True
        except Exception as e:
            write_debug(f"Could not close Excel processes: {str(e)}")
            return False
    except Exception as e:
        write_debug(f"Error checking file access: {str(e)}")
        return False

def initialize_excel():
    """Initialiserer Excel med robust fejlhåndtering og COM-objekt håndtering."""
    try:
        # Luk eventuelle eksisterende Excel processer
        os.system("taskkill /f /im excel.exe")
        time.sleep(2)  # Vent på at processerne lukkes
        
        # Initialiser COM objekter i den rigtige tråd
        pythoncom.CoInitialize()
        
        # Start Excel med retry
        max_retries = 3
        last_error = None
        
        for attempt in range(max_retries):
            try:
                excel = Dispatch("Excel.Application")
                excel.DisplayAlerts = False
                excel.Visible = False
                excel.EnableEvents = False
                excel.Interactive = False  # Deaktiver bruger interaktion
                write_debug("Successfully initialized Excel application")
                return excel
            except Exception as e:
                last_error = e
                write_debug(f"Attempt {attempt + 1} failed to initialize Excel: {str(e)}")
                time.sleep(2)
                try:
                    os.system("taskkill /f /im excel.exe")
                    time.sleep(2)
                except:
                    pass
        
        raise Exception(f"Could not initialize Excel after {max_retries} attempts. Last error: {str(last_error)}")
        
    except Exception as e:
        write_debug(f"Failed to initialize Excel: {str(e)}")
        raise

def cleanup_excel(excel=None, workbook=None, temp_file=None):
    """Sikrer proper cleanup af Excel ressourcer og COM objekter."""
    try:
        if workbook:
            try:
                workbook.Close(SaveChanges=False)
            except:
                pass
            finally:
                del workbook
        
        if excel:
            try:
                excel.Quit()
            except:
                pass
            finally:
                del excel
        
        # Frigør COM objekter
        pythoncom.CoUninitialize()
        
        # Vent og luk eventuelle hængende processer
        time.sleep(1)
        os.system("taskkill /f /im excel.exe")
        time.sleep(1)
        
        # Forsøg at slette temp fil efter Excel er lukket
        if temp_file and os.path.exists(temp_file):
            max_attempts = 3
            for attempt in range(max_attempts):
                try:
                    os.remove(temp_file)
                    write_debug(f"Successfully deleted temporary file on attempt {attempt + 1}")
                    break
                except Exception as e:
                    write_debug(f"Attempt {attempt + 1} to delete temp file failed: {str(e)}")
                    if attempt < max_attempts - 1:
                        time.sleep(2)
                    else:
                        write_debug("Could not delete temporary file after all attempts")
        
    except Exception as e:
        write_debug(f"Error during Excel cleanup: {str(e)}")

def create_temp_file(original_file, dest_dir):
    """Opretter en midlertidig fil med sikker håndtering af rettigheder."""
    try:
        # Generer unikt filnavn med timestamp og random del
        import uuid
        timestamp = time.strftime("%Y%m%d_%H%M%S")
        unique_id = str(uuid.uuid4())[:8]
        temp_filename = f"temp_{timestamp}_{unique_id}.xlsx"
        temp_path = os.path.join(dest_dir, temp_filename)
        
        write_debug(f"Attempting to create temporary file: {temp_path}")
        
        # Sikr at destinationsmappen har de rette rettigheder
        try:
            os.chmod(dest_dir, 0o777)  # Giv fulde rettigheder til mappen
            write_debug("Set full permissions on destination directory")
        except Exception as e:
            write_debug(f"Warning: Could not set directory permissions: {str(e)}")
        
        # Kopier filen med explicit rettigheder
        shutil.copy2(original_file, temp_path)
        try:
            os.chmod(temp_path, 0o666)  # Giv læse/skrive rettigheder til filen
            write_debug("Set read/write permissions on temporary file")
        except Exception as e:
            write_debug(f"Warning: Could not set file permissions: {str(e)}")
            
        write_debug(f"Successfully created temporary file: {temp_path}")
        return temp_path
        
    except Exception as e:
        write_debug(f"Error creating temporary file: {str(e)}")
        raise

def retry_operation(func, max_attempts=3, delay=1):
    """Implementerer retry mekanisme for filoperationer."""
    for attempt in range(max_attempts):
        try:
            return func()
        except Exception as e:
            if attempt == max_attempts - 1:
                raise
            write_debug(f"Attempt {attempt + 1} failed: {str(e)}. Retrying...")
            time.sleep(delay)

def process_excel(file_path):
    """Hovedfunktion der håndterer Excel-filen."""
    excel = None
    wb = None
    temp_path = None
    categorized_data = None
    
    try:
        write_debug(f"Starting process_excel with file: {file_path}")
        
        # Verificer at kildefilen eksisterer
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"Kildefilen findes ikke: {file_path}")
        
        # Opret filnavne og stier
        file_name = os.path.basename(file_path)
        part_number = file_name.split(" - ")[0]
        dest_dir = os.path.join(os.path.dirname(file_path), part_number)
        write_debug(f"Destination directory will be: {dest_dir}")
        
        # Opret destinationsmappe med retry
        def create_dir():
            os.makedirs(dest_dir, exist_ok=True)
        retry_operation(create_dir)
        write_debug(f"Successfully created destination directory: {dest_dir}")
        
        # Opret temp fil med retry og timestamp
        timestamp = time.strftime("%Y%m%d_%H%M%S")
        temp_path = os.path.join(dest_dir, f"temp_{timestamp}.xlsx")
        shutil.copy2(file_path, temp_path)
        write_debug(f"Created temporary file: {temp_path}")
        
        # Initialiser Excel og åbn workbook
        excel = initialize_excel()
        
        try:
            # Åbn workbook med retry
            def open_workbook():
                nonlocal wb
                wb = excel.Workbooks.Open(temp_path)
                sheet = wb.Sheets(1)
                sheet.Name = "BOM (Raw)"
                write_debug("Successfully opened workbook and renamed first sheet")
            retry_operation(open_workbook)
            
            # Valider og indlæs data med pandas
            df = pd.read_excel(temp_path, dtype=str)
            write_debug("Successfully loaded Excel file with pandas")
            
            # Kategoriser data
            categorized_data = categorize_data(df)
            write_debug(f"Data categorized into {len(categorized_data)} categories")
            
            # Gem ændringer og luk workbook før vi flytter filen
            if wb:
                wb.Save()
                wb.Close(SaveChanges=True)
                wb = None
            
            # Luk Excel helt
            if excel:
                excel.Quit()
                excel = None
            
            # Frigør COM objekter før filoperationer
            pythoncom.CoUninitialize()
            
            # Flyt temp fil til endelig destination
            final_path = os.path.join(dest_dir, f"{part_number} - BOM.xlsx")
            if os.path.exists(final_path):
                os.remove(final_path)
            os.rename(temp_path, final_path)
            write_debug("Successfully moved temporary file to final location")
            
            return final_path, categorized_data
            
        except Exception as e:
            write_debug(f"ERROR during Excel processing: {str(e)}")
            raise
        finally:
            cleanup_excel(excel, wb, temp_path)
            
    except Exception as e:
        write_debug(f"ERROR during Excel processing: {str(e)}")
        cleanup_excel(excel, wb, temp_path)
        messagebox.showerror("Error", f"Der opstod en fejl under behandling af Excel-filen:\n{str(e)}")
        return None, None

def newRev(name, files):
    """Finder den seneste revision ud fra filnavnet."""
    latest_files = {}
    for ext in [".pdf", ".dwg"]:
        relevant_files = [f for f in files if f.endswith(ext)]
        if relevant_files:
            latest_files[ext] = max(relevant_files, key=lambda f: f[-5:-4] if f[-5:-4].isalpha() else 'A')
    return latest_files

def scan_directory_task(entry, pdf_source):
    """Scanner undermapper for PDF- og DWG-filer."""
    if entry.is_dir(follow_symlinks=False):
        file_paths = []
        for sub_entry in os.scandir(entry.path):
            file_paths.extend(scan_directory_task(sub_entry, pdf_source))
        return file_paths
    elif entry.is_file(follow_symlinks=False) and (entry.name.endswith(".pdf") or entry.name.endswith(".dwg")):
        return [entry.path]
    return []

def scan_directory_concurrent(pdf_source):
    """Bruger multithreading til at scanne hele PDF-kataloget."""
    file_paths = []
    with ThreadPoolExecutor() as executor:
        futures = [executor.submit(scan_directory_task, entry, pdf_source) for entry in os.scandir(pdf_source)]
        for future in futures:
            file_paths.extend(future.result())
    return file_paths

def copy_pdf_dwg_files(dest_path, categorized_data, pdf_source):
    """Kopierer filer med retry mekanisme og caching."""
    try:
        # Brug cache til at få filer
        files = scan_directory_concurrent(pdf_source)
        
        # Tjek hver kategori for filer der skal kopieres
        for category, data in categorized_data.items():
            has_files = False
            # Tjek først om der er nogen filer at kopiere til denne kategori
            for _, row in data.iterrows():
                part_number = str(row.iloc[1])  # Kolonne 2 (Part Number)
                matching_files = [f for f in files if os.path.basename(f).startswith(part_number)]
                if matching_files:
                    has_files = True
                    break
            
            # Kun opret mappen og kopier filer hvis der faktisk er filer til denne kategori
            if has_files:
                category_path = os.path.join(dest_path, category)
                
                # Opret mappe med retry
                def create_dir():
                    os.makedirs(category_path, exist_ok=True)
                retry_operation(create_dir)
                
                for _, row in data.iterrows():
                    part_number = str(row.iloc[1])
                    matching_files = [f for f in files if os.path.basename(f).startswith(part_number)]
                    
                    if matching_files:
                        latest_files = newRev(part_number, matching_files)
                        for file in latest_files.values():
                            dest_file = os.path.join(category_path, os.path.basename(file))
                            
                            # Kopier fil med retry
                            def copy_file():
                                shutil.copy2(file, dest_file)
                            retry_operation(copy_file)
                            
                            # Log forskellige REV
                            if file.endswith('.pdf') and file.replace('.pdf', '.dwg') in latest_files:
                                write_debug(f"Different REV found for {part_number}: PDF={file}, DWG={file.replace('.pdf', '.dwg')}")
    
    except Exception as e:
        write_debug(f"ERROR during file copying: {str(e)}")
        raise

def create_gui():
    """Opretter GUI vinduet med alle påkrævede elementer."""
    root = Tk()
    root.attributes('-topmost', True)
    root.title("ExcelCopyBOM")
    
    # Main frame
    main_frame = ttk.Frame(root, padding="10")
    main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
    
    # File selection
    ttk.Label(main_frame, text="Open Excel BOM List:").grid(row=0, column=0, sticky=tk.W)
    bom_path = tk.StringVar()
    ttk.Entry(main_frame, textvariable=bom_path, width=50).grid(row=0, column=1, padx=5)
    ttk.Button(main_frame, text="Browse", command=lambda: browse_file(bom_path)).grid(row=0, column=2)
    
    ttk.Label(main_frame, text="Previous Drawing Package BOM List:").grid(row=1, column=0, sticky=tk.W)
    prev_path = tk.StringVar()
    ttk.Entry(main_frame, textvariable=prev_path, width=50).grid(row=1, column=1, padx=5)
    ttk.Button(main_frame, text="Browse", command=lambda: browse_file(prev_path)).grid(row=1, column=2)
    
    # PDF source directory
    ttk.Label(main_frame, text="PDF/DWG Source Directory:").grid(row=2, column=0, sticky=tk.W)
    pdf_source = tk.StringVar(value=r'C:\Coding\Python\ExcelCopyBOM\Files')
    ttk.Entry(main_frame, textvariable=pdf_source, width=50).grid(row=2, column=1, padx=5)
    ttk.Button(main_frame, text="Browse", command=lambda: browse_directory(pdf_source)).grid(row=2, column=2)
    
    # Checkboxes
    include_equipment = tk.BooleanVar()
    ttk.Checkbutton(main_frame, text="Include Equipment, Valve, Instrument", 
                    variable=include_equipment).grid(row=3, column=0, columnspan=3, sticky=tk.W)
    
    find_rev = tk.BooleanVar()
    ttk.Checkbutton(main_frame, text="Find REV files before date", 
                    variable=find_rev).grid(row=4, column=0, columnspan=3, sticky=tk.W)
    
    include_data = tk.BooleanVar()
    data_check = ttk.Checkbutton(main_frame, text="Include Data Sheet", 
                                variable=include_data)
    data_check.grid(row=5, column=0, columnspan=3, sticky=tk.W)
    
    # Date picker (initially disabled)
    date_var = tk.StringVar()
    date_entry = ttk.Entry(main_frame, textvariable=date_var, width=20)
    date_entry.grid(row=6, column=0, columnspan=3, sticky=tk.W)
    date_entry.state(['disabled'])
    
    # Progress bar
    progress = ttk.Progressbar(main_frame, mode='determinate', maximum=100)
    progress.grid(row=7, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=10)
    
    # Status label
    status_var = tk.StringVar(value="Ready")
    status_label = ttk.Label(main_frame, textvariable=status_var)
    status_label.grid(row=8, column=0, columnspan=3)
    
    # Start button
    ttk.Button(main_frame, text="Start", command=lambda: start_processing(
        bom_path.get(), prev_path.get(), include_equipment.get(), 
        find_rev.get(), include_data.get(), date_var.get(),
        progress, status_var, pdf_source.get()
    )).grid(row=9, column=0, columnspan=3, pady=10)
    
    # Bind checkbox events
    def on_equipment_change(*args):
        data_check.state(['disabled'] if not include_equipment.get() else ['!disabled'])
        if not include_equipment.get():
            include_data.set(False)
    
    def on_rev_change(*args):
        date_entry.state(['!disabled'] if find_rev.get() else ['disabled'])
    
    include_equipment.trace('w', on_equipment_change)
    find_rev.trace('w', on_rev_change)
    
    return root

def browse_file(path_var):
    """Åbner filvælger dialog og opdaterer path variabel."""
    filename = filedialog.askopenfilename(
        title="Select Excel file",
        filetypes=[("Excel files", "*.xlsx *.xls")]
    )
    if filename:
        path_var.set(filename)

def browse_directory(path_var):
    """Åbner mappe-vælger dialog og opdaterer path variabel."""
    directory = filedialog.askdirectory(
        title="Select PDF/DWG source directory"
    )
    if directory:
        path_var.set(directory)

def start_processing(bom_path, prev_path, include_equipment, find_rev, 
                    include_data, date, progress, status_var, pdf_source):
    """Starter processeringen af BOM filen."""
    if not bom_path:
        messagebox.showerror("Error", "Please select an Excel BOM file")
        return
    
    if not pdf_source:
        messagebox.showerror("Error", "Please select PDF/DWG source directory")
        return
    
    try:
        # Start processing
        status_var.set("Processing...")
        progress['value'] = 0
        
        # Process Excel file
        processed_file, categorized_data = process_excel(bom_path)
        progress['value'] = 30
        
        if processed_file is None:
            status_var.set("Error processing Excel file")
            return
            
        write_debug(f"Using PDF source directory: {pdf_source}")
            
        # Copy files
        status_var.set("Copying files...")
        if categorized_data:
            copy_pdf_dwg_files(os.path.dirname(processed_file), categorized_data, pdf_source)
        progress['value'] = 60
        
        # Compare if previous file selected
        if prev_path:
            status_var.set("Comparing with previous version...")
            compare_with_previous(processed_file, prev_path)
            progress['value'] = 90
            
        status_var.set("Complete!")
        progress['value'] = 100
        
        # Show completion dialog
        show_completion_dialog(processed_file)
        
    except Exception as e:
        status_var.set(f"Error: {str(e)}")
        messagebox.showerror("Error", str(e))

def show_completion_dialog(processed_file):
    """Viser dialog med resultater og mulighed for at gemme log."""
    log_text = read_debug_log()
    dialog = tk.Toplevel()
    dialog.title("Processing Complete")
    
    # Opret text widget med scrollbar
    frame = ttk.Frame(dialog)
    frame.pack(padx=10, pady=10, fill=tk.BOTH, expand=True)
    
    scrollbar = ttk.Scrollbar(frame)
    scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
    
    text = tk.Text(frame, height=20, width=60, yscrollcommand=scrollbar.set)
    text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
    scrollbar.config(command=text.yview)
    
    # Indsæt log tekst og gør read-only
    text.insert('1.0', log_text)
    text.configure(state='disabled')  # Brug configure i stedet for state
    
    # Knapper
    button_frame = ttk.Frame(dialog)
    button_frame.pack(pady=5)
    
    ttk.Button(button_frame, text="Save Log", 
               command=lambda: save_log(log_text, processed_file)).pack(side=tk.LEFT, padx=5)
    ttk.Button(button_frame, text="Open Folder", 
               command=lambda: open_folder(processed_file)).pack(side=tk.LEFT, padx=5)
    ttk.Button(button_frame, text="OK", 
               command=dialog.destroy).pack(side=tk.LEFT, padx=5)
    
    # Centrér dialogen på skærmen
    dialog.transient()
    dialog.grab_set()
    dialog.update_idletasks()
    
    # Beregn position
    screen_width = dialog.winfo_screenwidth()
    screen_height = dialog.winfo_screenheight()
    width = dialog.winfo_width()
    height = dialog.winfo_height()
    x = (screen_width - width) // 2
    y = (screen_height - height) // 2
    dialog.geometry(f'+{x}+{y}')

def read_debug_log():
    """Læser indholdet af Debug.txt filen."""
    debug_file = os.path.join(os.path.dirname(__file__), "Debug.txt")
    try:
        with open(debug_file, 'r', encoding='utf-8') as f:
            return f.read()
    except Exception as e:
        return f"Could not read debug log: {str(e)}"

def save_log(log_text, processed_file):
    """Gemmer log til en fil."""
    try:
        log_file = os.path.join(os.path.dirname(processed_file), "process_log.txt")
        with open(log_file, 'w', encoding='utf-8') as f:
            f.write(log_text)
        messagebox.showinfo("Success", f"Log saved to: {log_file}")
    except Exception as e:
        messagebox.showerror("Error", f"Could not save log: {str(e)}")

def open_folder(file_path):
    """Åbner Windows Stifinder i den angivne mappe."""
    folder_path = os.path.dirname(file_path)
    os.startfile(folder_path)

class ExcelHandler:
    """Håndterer Excel operationer med både openpyxl og win32com."""
    
    def __init__(self):
        self.excel = None
        self.wb = None
        self.image_map = {}
        
    def capture_image_info(self, workbook):
        """Gemmer information om alle billeder i workbook."""
        try:
            for sheet in workbook.Sheets:
                sheet_images = []
                for shape in sheet.Shapes:
                    image_info = {
                        'name': shape.Name,
                        'left': shape.Left,
                        'top': shape.Top,
                        'width': shape.Width,
                        'height': shape.Height,
                        'row': shape.TopLeftCell.Row,
                        'column': shape.TopLeftCell.Column,
                        'sheet': sheet.Name,
                        'relative_row': shape.TopLeftCell.Row - shape.TopLeftCell.Row,
                        'relative_col': shape.TopLeftCell.Column - shape.TopLeftCell.Column
                    }
                    sheet_images.append(image_info)
                self.image_map[sheet.Name] = sheet_images
            write_debug(f"Captured information for {len(self.image_map)} sheets with images")
        except Exception as e:
            write_debug(f"Error capturing image info: {str(e)}")
            raise
    
    def process_excel_with_images(self, file_path):
        """Hovedfunktion der håndterer Excel med billeder."""
        try:
            write_debug(f"Starting Excel processing with images for file: {file_path}")
            
            # Initialiser COM objekter
            pythoncom.CoInitialize()
            
            # Start Excel og åbn workbook
            self.excel = Dispatch("Excel.Application")
            self.excel.Visible = False
            self.excel.DisplayAlerts = False
            
            # Åbn original fil og gem billede information
            self.wb = self.excel.Workbooks.Open(file_path)
            self.capture_image_info(self.wb)
            
            # Luk Excel midlertidigt
            self.wb.Close(SaveChanges=False)
            self.excel.Quit()
            pythoncom.CoUninitialize()
            
            # Brug openpyxl til data manipulation
            wb_openpyxl = load_workbook(file_path)
            
            # Manipuler data som nødvendigt
            self._process_data(wb_openpyxl)
            
            # Gem ændringer
            temp_path = file_path.replace(".xlsx", "_temp.xlsx")
            wb_openpyxl.save(temp_path)
            
            # Genåbn med Excel for at gendanne billeder
            pythoncom.CoInitialize()
            self.excel = Dispatch("Excel.Application")
            self.excel.Visible = False
            self.wb = self.excel.Workbooks.Open(temp_path)
            
            # Gendan billeder i deres korrekte positioner
            self._restore_images()
            
            # Gem endelige ændringer
            self.wb.SaveAs(file_path)
            return file_path
            
        except Exception as e:
            write_debug(f"Error in process_excel_with_images: {str(e)}")
            raise
        finally:
            if self.wb:
                try:
                    self.wb.Close(SaveChanges=False)
                except:
                    pass
            if self.excel:
                try:
                    self.excel.Quit()
                except:
                    pass
            pythoncom.CoUninitialize()
    
    def _process_data(self, workbook):
        """Udfører data manipulation med openpyxl."""
        try:
            # Indstil kolonnebredder ifølge project.rules
            for sheet in workbook.worksheets:
                # Sæt standard kolonnebredder
                for col in range(1, sheet.max_column + 1):
                    column_letter = get_column_letter(col)
                    sheet.column_dimensions[column_letter].width = 15  # Standard bredde
                
                # Sæt specifik rækkehøjde
                sheet.row_dimensions[1].height = 20  # Header højde
                for row in range(2, sheet.max_row + 1):
                    sheet.row_dimensions[row].height = 91  # Data række højde
                
                # Frys første række
                sheet.freeze_panes = 'A2'
                
                # Tilføj filter til første række
                sheet.auto_filter.ref = f"A1:{get_column_letter(sheet.max_column)}1"
                
            write_debug("Successfully processed data with openpyxl")
            
        except Exception as e:
            write_debug(f"Error in _process_data: {str(e)}")
            raise
    
    def _restore_images(self):
        """Gendanner billeder i deres korrekte positioner."""
        try:
            for sheet_name, images in self.image_map.items():
                sheet = self.wb.Sheets(sheet_name)
                for image_info in images:
                    shape = sheet.Shapes(image_info['name'])
                    # Gendan original position
                    shape.Top = image_info['top']
                    shape.Left = image_info['left']
                    shape.Width = image_info['width']
                    shape.Height = image_info['height']
            write_debug("Successfully restored all images")
        except Exception as e:
            write_debug(f"Error restoring images: {str(e)}")
            raise

def main():
    root = create_gui()
    root.mainloop()

if __name__ == "__main__":
    main()


