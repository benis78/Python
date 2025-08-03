import os
import shutil
import pandas as pd
import openpyxl
import time
import sqlite3
from datetime import datetime, timedelta
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font
from tkinter import Tk, filedialog, messagebox, ttk
import tkinter as tk
from win32com.client import Dispatch
from concurrent.futures import ThreadPoolExecutor
import subprocess
import threading

# Konfiguration
NETVAERKSDREV = r'\\192.168.170.18\Drawings'
DATABASE_PATH = os.path.join(NETVAERKSDREV, "file_index.db")
INDEXER_PATH = os.path.join(NETVAERKSDREV, "file_indexer.exe")

class IndexingStatus:
    def __init__(self):
        self.window = tk.Tk()
        self.window.title("Indeksering Status")
        self.window.geometry("400x150")
        
        # Centrer vinduet på skærmen
        screen_width = self.window.winfo_screenwidth()
        screen_height = self.window.winfo_screenheight()
        x = (screen_width - 400) // 2
        y = (screen_height - 150) // 2
        self.window.geometry(f"400x150+{x}+{y}")
        
        # Status label
        self.status_label = tk.Label(self.window, text="Indekserer filer...", font=("Arial", 12))
        self.status_label.pack(pady=10)
        
        # Progress bar
        self.progress = ttk.Progressbar(self.window, length=300, mode='indeterminate')
        self.progress.pack(pady=10)
        
        # Detalje label
        self.detail_label = tk.Label(self.window, text="Venter på opdatering...", font=("Arial", 10))
        self.detail_label.pack(pady=5)
        
        self.progress.start(10)  # Start progress bar animation
        
        # Flag til at holde styr på om indexer kører
        self.indexing = True
        
        # Start thread til at tjekke database status
        self.check_thread = threading.Thread(target=self.check_indexing_status)
        self.check_thread.daemon = True
        self.check_thread.start()
    
    def check_indexing_status(self):
        while self.indexing:
            try:
                conn = sqlite3.connect(DATABASE_PATH)
                cursor = conn.cursor()
                cursor.execute("SELECT value FROM metadata WHERE key = 'last_scan_time'")
                result = cursor.fetchone()
                conn.close()
                
                if result:
                    last_scan_time = float(result[0])
                    last_scan_date = datetime.fromtimestamp(last_scan_time)
                    current_time = datetime.now()
                    time_diff = (current_time - last_scan_date).total_seconds()
                    
                    if time_diff < 60:  # Hvis sidste scan er mindre end 1 minut gammelt
                        self.indexing = False
                        self.window.after(0, self.finish_indexing)
                        break
                    
                    self.detail_label.config(text=f"Sidste opdatering: {last_scan_date.strftime('%H:%M:%S')}")
            except Exception as e:
                self.detail_label.config(text=f"Tjekker status...")
            
            time.sleep(2)  # Vent 2 sekunder mellem hver tjek
    
    def finish_indexing(self):
        self.progress.stop()
        self.status_label.config(text="Indeksering færdig!")
        self.detail_label.config(text="Du kan nu lukke dette vindue")
        self.window.after(2000, self.window.destroy)  # Luk vinduet efter 2 sekunder

def choose_file():
    """Lader brugeren vælge en Excel-fil og returnerer dens sti."""
    root = Tk()
    root.withdraw()
    # file_path = filedialog.askopenfilename(title="Open Excel BOM file",
    #                                        filetypes=[("Excel files", ".xlsx .xls")],
    #                                        initialdir="C:\\Working Folder\\Designs\\5-Projects")
    file_path = "C:\\Coding\\Python\\ExcelCopyBOM\\4003-02.1-A01-- - BOM.xlsx"
    if not file_path:
        os._exit(1)
    return file_path

def create_copy(file_path):
    """Opretter en kopi af den originale Excel-fil."""
    dest_path = os.path.join(os.path.dirname(file_path), "BOM_Copy.xlsx")
    shutil.copy(file_path, dest_path)
    return dest_path

def newRev(name, files):
    """Finder den seneste revision ud fra filnavnet."""
    latest_files = {}
    for ext in [".pdf", ".dwg"]:
        relevant_files = [f for f in files if f.endswith(ext)]
        if relevant_files:
            latest_files[ext] = max(relevant_files, key=lambda f: f[-5:-4] if f[-5:-4].isalpha() else 'A')
    return latest_files

def get_latest_revision(files):
    """Finder den seneste revision blandt de fundne filer"""
    if not files:
        return None
    # Sorter efter revisionsbogstav (sidste bogstav før filtypen)
    latest = max(files, key=lambda f: f[1][-5:-4] if f[1][-5:-4].isalpha() else 'A')
    return latest[0]  # Returner stien til den nyeste revision

def load_categories():
    """Indlæser kategorier fra en ekstern .txt-fil."""
    categories = {}
    categories_file = os.path.join(os.path.dirname(__file__), "categories.txt")  # Sikrer, at filen hentes fra script-mappen

    if not os.path.exists(categories_file):
        messagebox.showerror("ERROR", f"Categories file not found: {categories_file}")
        return categories  # Returnerer en tom liste, så programmet ikke fejler

        with open(categories_file, "r") as file:
            for line in file:
                parts = line.strip().split("=")
                if len(parts) == 2:
                    categories[parts[0].strip()] = parts[1].strip()

    return categories

def categorize_data(df):
    """Opdeler data i kategorier baseret på Part Number (kolonne 2) og returnerer en dict med kategorier."""
    categories = load_categories()
    categorized_data = {}
    
    for _, row in df.iterrows():
        part_number = str(row.iloc[1])  # Kolonne 2 (Part Number)
        for prefix, cat_name in categories.items():
            if part_number.startswith(prefix):
                if cat_name not in categorized_data:
                    categorized_data[cat_name] = pd.DataFrame(columns=df.columns)
                categorized_data[cat_name] = pd.concat([categorized_data[cat_name], row.to_frame().T], ignore_index=True)
                break
        else:
            if "Other Items" not in categorized_data:
                categorized_data["Other Items"] = pd.DataFrame(columns=df.columns)
            categorized_data["Other Items"] = pd.concat([categorized_data["Other Items"], row.to_frame().T], ignore_index=True)
    
    return categorized_data

def process_excel(file_path):
    """Behandler Excel-filen, laver en kopi, og opdeler data i faner."""
    copy_path = create_copy(file_path)
    wb = openpyxl.load_workbook(copy_path)
    sheet = wb.active
    sheet.title = "BOM (Raw)"
    
    df = pd.read_excel(copy_path, sheet_name=0, dtype=str)
    categorized_data = categorize_data(df)
    
    for sheet_name, data in categorized_data.items():
        ws = wb.create_sheet(title=sheet_name)
        for col_num, col_name in enumerate(data.columns, 1):
            ws.cell(row=1, column=col_num, value=col_name)
        for row_num, row in data.iterrows():
            for col_num, value in enumerate(row, 1):
                ws.cell(row=row_num + 2, column=col_num, value=value)
    
    wb.save(copy_path)
    return copy_path, categorized_data

def search_files(part_number, file_type=None):
    """
    Søger i databasen efter filer der matcher part_number
    :param part_number: Partnummer at søge efter
    :param file_type: Filtype filter (f.eks. '.pdf' eller '.dwg')
    :return: Liste af fundne filer (sti, filnavn)
    """
    try:
        conn = sqlite3.connect(DATABASE_PATH)
        cursor = conn.cursor()
        
        # Fjern eventuelle mellemrum og konverter til lowercase for bedre søgning
        part_number = part_number.strip().lower()
        
        # Søg efter filer der starter med part_number
        if file_type:
            cursor.execute("""
                SELECT path, filename
                FROM files 
                WHERE LOWER(filename) LIKE ? AND file_type = ?
            """, (f"{part_number}%", file_type.lower()))
        else:
            cursor.execute("""
                SELECT path, filename
                FROM files 
                WHERE LOWER(filename) LIKE ?
            """, (f"{part_number}%",))
            
        results = cursor.fetchall()
        conn.close()
        
        # Debug print
        print(f"Søgeresultater for {part_number}: {len(results)} filer fundet")
        for path, filename in results:
            print(f"- {os.path.join(path, filename)}")
        
        return [(os.path.join(path, filename), filename) for path, filename in results]
    except Exception as e:
        print(f"Fejl ved søgning i database: {e}")
    return []

def copy_pdf_dwg_files(dest_path, categorized_data):
    """Kopierer den nyeste revision af PDF- og DWG-filer til relevante mapper baseret på Part Number"""
    total_files = 0
    copied_files = 0
    
    for category, data in categorized_data.items():
        category_path = os.path.join(dest_path, category)
        os.makedirs(category_path, exist_ok=True)
        print(f"\nBehandler kategori: {category}")
        
        for _, row in data.iterrows():
            part_number = str(row.iloc[1]).strip()  # Kolonne 2 (Part Number)
            if not part_number:  # Spring over tomme partnumre
                continue
                
            total_files += 1
            print(f"\nSøger efter filer for: {part_number}")
            
            # Søg efter PDF og DWG filer
            for ext in [".pdf", ".dwg"]:
                try:
                    files = search_files(part_number, ext)
                    if files:
                        print(f"Fandt {len(files)} {ext} filer for {part_number}")
                        latest = get_latest_revision(files)
                        if latest:
                            try:
                                source_path = latest
                                if os.path.exists(source_path):
                                    dest_file = os.path.join(category_path, os.path.basename(source_path))
                                    print(f"Kopierer {source_path} til {dest_file}")
                                    shutil.copy2(source_path, dest_file)
                                    copied_files += 1
                                    print(f"Kopiering af {os.path.basename(source_path)} gennemført")
                                else:
                                    print(f"ADVARSEL: Kildefil findes ikke: {source_path}")
                            except Exception as e:
                                print(f"Fejl ved kopiering af fil {source_path}: {e}")
                    else:
                        print(f"Ingen {ext} filer fundet for {part_number}")
                except Exception as e:
                    print(f"Fejl ved søgning efter {ext} filer for {part_number}: {e}")
    
    print(f"\nTotal statistik:")
    print(f"Behandlede partnumre: {total_files}")
    print(f"Kopierede filer: {copied_files}")
    return total_files, copied_files

def check_database_age():
    """Tjekker alderen på databasen og opdaterer hvis nødvendigt"""
    try:
        conn = sqlite3.connect(DATABASE_PATH)
        cursor = conn.cursor()
        cursor.execute("SELECT value FROM metadata WHERE key = 'last_scan_time'")
        result = cursor.fetchone()
        conn.close()
        
        if result:
            last_scan_time = float(result[0])
            last_scan_date = datetime.fromtimestamp(last_scan_time)
            current_time = datetime.now()
            age_minutes = (current_time - last_scan_date).total_seconds() / 60
            
            # Hvis databasen er ældre end 30 minutter
            if age_minutes > 30:
                print(f"Indeks er {int(age_minutes)} minutter gammelt. Opdaterer...")
                if os.path.exists(INDEXER_PATH):
                    try:
                        # Start indexer som en separat proces
                        process = subprocess.Popen([INDEXER_PATH], 
                                                 stdout=subprocess.PIPE,
                                                 stderr=subprocess.PIPE,
                                                 creationflags=subprocess.CREATE_NO_WINDOW)
                        
                        # Vis status vindue
                        status = IndexingStatus()
                        status.window.mainloop()
                        
                        # Vent på at processen er færdig
                        process.wait()
                        
                        if process.returncode == 0:
                            print("Indeksering færdig")
                            return True
                        else:
                            messagebox.showerror("Fejl", "file_indexer.exe fejlede under kørsel")
                            return False
                    except Exception as e:
                        messagebox.showerror("Fejl", f"Kunne ikke starte file_indexer.exe:\n{e}")
                        return False
                else:
                    messagebox.showerror("Fejl", f"Kan ikke finde file_indexer.exe på stien:\n{INDEXER_PATH}")
                    return False
            else:
                print(f"Indeks er opdateret ({int(age_minutes)} minutter gammelt)")
                return True
            
        return True
    except Exception as e:
        messagebox.showerror("Fejl", f"Fejl ved tjek af database alder:\n{e}")
        return False

def main():
    start_time = time.time()
    
    # Tjek om databasen er tilgængelig
    if not os.path.exists(DATABASE_PATH):
        messagebox.showerror("Error", f"Kan ikke finde databasen: {DATABASE_PATH}")
        return
    
    # Tjek databasens alder
    if not check_database_age():
        return  # Afslut hvis brugeren vælger at opdatere indekset
    
    file_path = choose_file()
    processed_file, categorized_data = process_excel(file_path)
    
    base_name = os.path.basename(file_path).replace(" - BOM.xlsx", "")
    dest_path = os.path.join(os.path.dirname(file_path), base_name)
    os.makedirs(dest_path, exist_ok=True)
    
    #pdf_source = r'\\192.168.170.18\drawings'
    #pdf_source = r'C:\Working Folder\Designs\5-Projects\4003 - Nurmo Bioenergia\Area 05 - Storage Area\Equipment\BOM\Test\Files'
    #pdf_source = r'C:\\Coding\\Python\\ExcelCopyBOM\\Files'
    #copy_pdf_dwg_files(dest_path, categorized_data, pdf_source)
    total_files, copied_files = copy_pdf_dwg_files(dest_path, categorized_data)
    
    duration = time.time() - start_time
    messagebox.showinfo("Success", 
                       f"BOM processing complete!\n"
                       f"Found {total_files} part numbers\n"
                       f"Copied {copied_files} files\n"
                       f"Files saved in: {dest_path}\n"
                       f"Time taken: {duration:.2f} seconds")

if __name__ == "__main__":
    main()


