# import os, openpyxl, shutil, re, time
# import pandas as pd
# import tkinter as tk
# from tkinter import filedialog, messagebox
# import win32com.client as win32
# from concurrent.futures import ThreadPoolExecutor

import os
import shutil
import re
import time
import openpyxl
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from win32com.client import Dispatch  # Importer kun Dispatch fra win32com.client
from concurrent.futures import ThreadPoolExecutor
from pandas import read_excel  # Importer kun read_excel fra pandas
import logging  # Tilføj logging
import sys
from pathlib import Path
from PyQt5.QtWidgets import QApplication
from gui import MainWindow

# Konfigurer logging
logging.basicConfig(
    filename='ExcelCopyBOM.log',
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)

class ExcelCopyBOMGUI:
    def __init__(self):
        self.root = tk.Tk()
        self.root.title("Excel Copy BOM")
        self.root.attributes('-topmost', True)
        
        # Centrér vinduet
        window_width = 500
        window_height = 200
        screen_width = self.root.winfo_screenwidth()
        screen_height = self.root.winfo_screenheight()
        x = (screen_width - window_width) // 2
        y = (screen_height - window_height) // 2
        self.root.geometry(f'{window_width}x{window_height}+{x}+{y}')
        
        # Excel fil vælger
        frame = ttk.Frame(self.root, padding="10")
        frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        ttk.Label(frame, text="Excel BOM File:").grid(row=0, column=0, sticky=tk.W)
        self.file_path_var = tk.StringVar()
        self.file_entry = ttk.Entry(frame, textvariable=self.file_path_var, width=50)
        self.file_entry.grid(row=0, column=1, padx=5)
        
        ttk.Button(frame, text="Browse", command=self.browse_file).grid(row=0, column=2)
        
        # DWG Checkbox
        self.include_dwg = tk.BooleanVar()
        ttk.Checkbutton(frame, text="Include DWG files", variable=self.include_dwg).grid(row=1, column=0, columnspan=3, pady=5)
        
        # Procesbar
        self.progress_var = tk.DoubleVar()
        self.progress = ttk.Progressbar(frame, length=400, mode='determinate', variable=self.progress_var)
        self.progress.grid(row=2, column=0, columnspan=3, pady=20)
        
        # Status label
        self.status_var = tk.StringVar(value="Ready to start...")
        self.status_label = ttk.Label(frame, textvariable=self.status_var)
        self.status_label.grid(row=3, column=0, columnspan=3)
        
        # Start knap
        self.start_button = ttk.Button(frame, text="Start Processing", command=self.start_processing)
        self.start_button.grid(row=4, column=0, columnspan=3, pady=10)

    def browse_file(self):
        file_path = filedialog.askopenfilename(
            title="OPEN EXCEL BOM FILE",
            filetypes=[('Excel files', '.xlsx .xls')],
            initialdir='C:\\Working Folder\\Designs\\5-Projects'
        )
        if file_path:
            self.file_path_var.set(file_path)

    def update_progress(self, value, message):
        self.progress_var.set(value)
        self.status_var.set(message)
        self.root.update()

    def start_processing(self):
        file_path = self.file_path_var.get()
        if not file_path:
            messagebox.showerror("Error", "Please select an Excel BOM file first.")
            return
            
        self.start_button.config(state='disabled')
        self.update_progress(0, "Starting process...")
        
        try:
            self.process_bom(file_path)
        except Exception as e:
            logging.error(f"Error during processing: {str(e)}", exc_info=True)
            messagebox.showerror("Error", f"An error occurred: {str(e)}")
            self.start_button.config(state='normal')

    def process_bom(self, file_path):
        # Eksisterende variabler
        efile = os.path.basename(file_path)
        ePath = os.path.dirname(file_path)
        eBOM = os.path.abspath(file_path)
        
        # Lav en backup af den originale fil
        backup_file = os.path.join(ePath, 'original_' + efile)
        shutil.copy2(file_path, backup_file)
        
        self.update_progress(10, "Checking file permissions...")
        
        # Tjek om kildefilen er skrivebeskyttet
        if not os.access(file_path, os.W_OK):
            messagebox.showerror("ERROR", f"The Excel file '{efile}' is not writeable check the file out of the vault and try again.")
            return

        self.update_progress(20, "Creating destination folder...")
        
        # Find positionerne af de første 3 '-' i filnavnet
        dash_positions = [pos for pos, char in enumerate(efile) if char == '-']

        # Opret mappenavn ved at tage alt indtil det tredje '-' plus det første tegn efter det tredje '-'
        if len(dash_positions) >= 3:
            # Find alt indtil det tredje '-'
            base_name = efile[:dash_positions[2] + 1]  # Inkluderer tredje '-'
            
            # Tag det fjerde tegn efter det tredje '-'
            if len(efile) > dash_positions[2] + 1:
                base_name += efile[dash_positions[2] + 1]  # Tilføj det første tegn efter tredje '-'

            new_folder_name = base_name
        else:
            # Hvis der ikke er nok '-', brug hele filnavnet
            new_folder_name = efile.split('.')[0]  # Brug filnavnet uden udvidelse

        # Definer destinationen for den nye mappe baseret på hvor Excel-filen ligger
        dest_path = os.path.join(ePath, new_folder_name)

        # Opret mappen, hvis den ikke allerede findes
        if not os.path.exists(dest_path):
            os.makedirs(dest_path)

        destPath=(os.path.abspath(dest_path))
        destDir=(os.path.basename)
        if dest_path == '':
            os._exit(1)    

        #efile='2205-401-001-D - Digester Area.xlsx'
        #destPath=('c:\\test\\drawings')
        #eBOM=(r'c:\\test\\'+efile)
        #source = (r'\\192.168.170.18\drawings')
        source = (r'C:\Coding\Python\ExcelCopyBOM\Files')
        #eSource=('c:\\test')

        start_time = time.time()

        def newRev(name, files):
            # Opdel filer efter filtype
            pdf_files = [f for f in files if f.endswith('.pdf')]
            dwg_files = [f for f in files if f.endswith('.dwg')]
            
            # Initialiser returværdier
            latestFiles = []
            revChar = '-'
            
            # Find seneste PDF
            if pdf_files:
                latestPdf = pdf_files[0]
                for rev in pdf_files:
                    revPos = rev.find(str(name)) + len(str(name)) + 1
                    revLetter = rev[revPos]
                    if revLetter > revChar:
                        latestPdf = rev
                        revChar = revLetter
                latestFiles.append(latestPdf)
            
            # Find seneste DWG med samme revision hvis DWG er aktiveret
            if self.include_dwg.get() and dwg_files:
                # Find DWG med samme revision som PDF
                matching_dwg = None
                for dwg in dwg_files:
                    revPos = dwg.find(str(name)) + len(str(name)) + 1
                    if dwg[revPos] == revChar:
                        matching_dwg = dwg
                        break
                
                # Hvis ingen DWG med samme revision findes, tag den seneste DWG
                if not matching_dwg:
                    matching_dwg = dwg_files[0]
                    for rev in dwg_files:
                        revPos = rev.find(str(name)) + len(str(name)) + 1
                        revLetter = rev[revPos]
                        if revLetter > revChar:
                            matching_dwg = rev
                            revChar = revLetter
                
                if matching_dwg:
                    latestFiles.append(matching_dwg)
            
            return latestFiles, revChar

        # Hjælpefunktion til at scanne individuelle entries i concurrent.futures
        def scan_directory_task(entry):
            if entry.is_dir(follow_symlinks=False):
                # Hvis det er en mappe, så gå videre og scan undermapperne
                file_paths = []
                for sub_entry in os.scandir(entry.path):
                    file_paths.extend(scan_directory_task(sub_entry))
                return file_paths
            elif entry.is_file(follow_symlinks=False):
                # Hvis det er en fil, sammenlign med targetFileNames og targetExtensions
                name = entry.name
                for targetName in targetFileNames:
                    for targetExtension in targetExtensions:
                        if name.startswith(targetName) and name.endswith(targetExtension):
                            if not name.endswith(excludeExtensions):
                                idx = -1
                                while True:
                                    try:
                                        idx = targetFileNames.index(targetName, idx + 1)
                                        fileDestinations[idx].append(entry.path)
                                    except ValueError:
                                        break
                return [entry.path]
            return []

        # Funktion til concurrent.futures scanning med os.scandir
        def scan_directory_concurrent(directory):
            with ThreadPoolExecutor() as executor:
                futures = [executor.submit(scan_directory_task, entry) for entry in os.scandir(directory)]
                for future in futures:
                    future.result()  # Vi skal vente på alle tråde at afslutte

        # Opret forbindelse til Excel via COM
        excel = Dispatch('Excel.Application')
        excel.Visible = False  # Kør Excel i baggrunden
        workbook = excel.Workbooks.Open(str(eBOM))
        sheet = workbook.Sheets('BOM')

        # Mønstre som skal fjernes i kolonne B
        patterns_to_remove = ['0000-700', '0000-701', '0000-702', '0000-704', '0000-705']
        pattern = re.compile(r'^(' + '|'.join(patterns_to_remove) + ')')

        # Få det samlede antal brugte rækker i arket ved at bruge Excel's egen metode
        max_row = sheet.UsedRange.Rows.Count

        # Iterer gennem rækkerne og fjern dem, der matcher mønsteret eller starter med tekst i kolonne B
        rows_to_delete = []
        for row in range(2, max_row + 1):  # Start fra anden række for at undgå headers
            cell_value = str(sheet.Cells(row, 2).Value)  # Kolonne B

            # Fjern rækker, hvis cellen matcher mønstre eller starter med bogstav
            if pattern.match(cell_value) or (cell_value and cell_value[0].isalpha()):  # Tjek om første tegn er et bogstav
                rows_to_delete.append(row)

        # Slet rækker (fra bunden op for at undgå skift i indeks)
        for row in reversed(rows_to_delete):
            sheet.Rows(row).Delete()
        efile_parts = efile.split('-')

        if len(efile_parts) >= 4:  # Sørg for, at der er mindst 4 dele
            # Indsæt en ny række i Excel på række 2
            sheet.Rows(2).Insert()

            # Indsæt filnavnet (op til tredje '-') i B2
            sheet.Cells(2, 2).Value = '-'.join(efile_parts[:3])

            # Indsæt tallet 0 i A2
            sheet.Cells(2, 1).Value = 0

            # Indsæt bogstavet efter tredje '-' i C2, hvis det findes
            if efile_parts[3]:  # Tjek om der er et bogstav efter tredje '-'
                sheet.Cells(2, 3).Value = efile_parts[3][0]  # Første tegn efter tredje '-'
            else:
                sheet.Cells(2, 3).Value = "-"  # Hvis ingen bogstav findes, indsæt '-'

            # Indsæt tallet tekst Top assembly i G2
            sheet.Cells(2, 7).Value = "Top Assembly"

            # Indsæt tallet 1 i J2
            sheet.Cells(2, 10).Value = 1
        # Gem og luk Excel-filen
        workbook.Save()
        workbook.Close(SaveChanges=True)
        excel.Quit()

        # os.chdir(ePath)
        wb=openpyxl.load_workbook(eBOM)
        sheet = wb['BOM']

        # get max row count
        max_row=sheet.max_row
        #Start row
        r=1      
        # get max column count
        max_column=sheet.max_column

        targetFileNames=[]
        for i in range(2,max_row+1):
          # iterate over all columns
          # get particular cell value
          cell_obj=sheet.cell(row=i,column=2)
          targetFileNames.append(cell_obj.value)
          #print (targetFileNames)

        # Definer filtyper baseret på checkbox
        targetExtensions = ['.pdf']
        if self.include_dwg.get():
            targetExtensions.append('.dwg')
            
        excludeExtensions= '_FOR REVIEW.pdf'

        fileDestinations = [[] for target in targetFileNames] # liste med samme længde som variablen targetFileNames, hvor hvert element til at starte med selv er en tom liste.
        #print (fileDestinations)
        # Værdi i variablen nu:
        # fileDestinations = [[], [], []]

        # for root, dirs, files in os.walk(source):           # Vi looper igennem vores directory
        #     if 'Old' in dirs:
        #         dirs.remove('Old')
        #     if 'old' in dirs:
        #         dirs.remove('old')
        #     if 'OLD' in dirs:
        #         dirs.remove('OLD')
        #     for name in files:                              # Vi looper igennem hver fil i den aktuelle mappe
        #         #Hver fil bliver så sammenlignet med hver værdi i targetFileNames med hver af de mulige extensions:
        #         for targetName in targetFileNames:
        #             for targetExtension in targetExtensions:
        #                 if name.startswith(targetName) and name.endswith(targetExtension): 
        #                     if not name.endswith(excludeExtensions):   
        #                     # Når vi finder et match, bliver stien til den aktuelle fil tilføjet på den passende position i variablen fileDestinations:
        #                         #fileDestinations[targetFileNames.index(targetName)]
        #                         #fileDestinations[targetFileNames.index(targetName)].append(os.path.join(root,name))
        #                         idx=-1
        #                         while True:
        #                             try:
        #                                 idx = targetFileNames.index(targetName, idx+1)
        #                                 fileDestinations[idx].append(os.path.join(root,name))
        #                             except ValueError:
        #                                 break
                                #fileDestinations[ti].append(os.path.join(root,name))
                                #print(fileDestinations, end= '\n')
        
        #print(targetName, end= '\n')
                        #print(name, end='\n') #Denne virker i for løkken og returere hele filnavnet men kun i løkken.

        if __name__ == "__main__":
            # Start scanning med concurrent.futures
            scan_directory_concurrent(source)

        #####################################

        data=read_excel(file_path,sheet_name='BOM',header=0)

        # Convert QTY To String
        data['QTY'] = data['QTY'].astype(str)
        # Convert string to int
        #qty=[int(i) for i in data['QTY']]
        qty = [int(ele) if ele.isdigit() else int(ele.rsplit(',', 1)[0]) for ele in data['QTY']] 

        # Convert ITEM to String
        data['Item']= data['Item'].astype(str)
        # Remove .0 in Item column
        data['Item']=[i.rsplit('.0', 1)[0] for i in data['Item']]
        # Get parent level
        parentlevels=[i.rsplit('.', 1)[0] for i in data['Item']]
        # Get parent level with a empty string as first level I then can use in my for loop at the end.
        pp=[i.rpartition('.')[0] for i in data['Item']]
        #print(parentlevels)

        # Put Items in a list called items
        items=[]
        for index, value in enumerate(data['Item']): 
            items.append(value)

        # Search for items in  parentlevels 
        ii=[items.index(i) for i in parentlevels]

        totalQTY=[]
        e=0
        for i in pp:
            if i =='':
                totalQTY.append(qty[e])
                #print(i)
            else:
                totalQTY.append(qty[e]*totalQTY[ii[e]])
            e+=1

        colQTY = data.columns.get_loc('QTY')

        #tqty=data['Total QTY']=totalQTY

        data.insert(colQTY+1,column='Total QTY',value=totalQTY)

        #####################################

        excelList=[[] for target in targetFileNames]  # Ændret til liste af lister
        revList=['' for target in targetFileNames]

        i=0
        for destination in fileDestinations:
            if len(destination) > 0:
                files, revChar = (newRev(targetFileNames[i], destination))
                excelList[i] = files  # Nu en liste af filer
                revList[i] = revChar 
            i+=1

        # set col to next
        col=max_column+1

        sheet.cell(row=1, column=col).value = 'Drawing'

        # if not dest_path == '':
        #     sheet.cell(row=1, column=col+1).value = 'Local Drawing'

        i = 2
        for files in excelList:
            if not files:  # Hvis der ikke er nogen filer
                sheet.cell(row=i, column=col).value = 'Not available'
            else:
                # Kopier alle filer til destinationsmappen
                for file in files:
                    shutil.copy2(file, destPath)
                
                # Find PDF-filen (hvis den findes)
                pdf_file = next((f for f in files if f.endswith('.pdf')), None)
                
                if pdf_file:
                    localPath = os.path.basename(pdf_file)
                    # Indsæt hyperlink kun for PDF
                    sheet.cell(row=i, column=col).value = f'=HYPERLINK("{localPath}","PDF")'
                    sheet.cell(row=i, column=col).style = 'Hyperlink'
                else:
                    sheet.cell(row=i, column=col).value = 'Not available'
            
            i += 1

        sheet.insert_cols(11,1)
        sheet.cell(row=1, column=11).value = 'Total QTY'

        i=2
        for elements in totalQTY:
            sheet.cell(row=i, column=11).value = elements
            i+=1

        # Get column number of REV header
        for col in sheet.iter_cols(1):
            for cell in col:
                if cell.value == "REV":
                    c_rev=(sheet.cell(row=1, column=cell.column).column)
        
        # insert revion letter in excel
        i=2
        for File in revList:
            sheet.cell(row=i, column=c_rev).value = (File)
            i+=1

        timestr = time.strftime("%Y%m%d-%H%M")

        # copy drawings to ZIP folder
        # if not dest_path=="":
            #  shutil.make_archive(destPath+'-'+'zip',destPath)
            # shutil.make_archive(destPath+'-'+'zip','zip')
        # os.rename('a.txt', 'b.kml')

        # save excel         
        wb.save(str(destPath)+'\\'+str(efile))

        # if not dest_path=="":
        #     shutil.make_archive(destPath+'-'+timestr,'zip',destPath)
        
        # Kopier den originale fil tilbage
        shutil.copy2(backup_file, file_path)
        # Slet backup filen
        os.remove(backup_file)
        
        duration=(time.time() - start_time)

        messagebox.showinfo(title='ExcelCopyBOM', message='Done\nTime: '+"{0:.2f}".format(duration)+' seconds')

        # print("--- %s seconds ---" % (time.time() - start_time))

        #Pyinstaller --onefile ExcelCopyBOM.py
        #PyInstaller --onefile --upx-dir=D:\OneDrive\Coding\upx-4.2.4-win64 ExcelCopyBOM.py

        # Når processen er færdig
        self.update_progress(100, "Process completed!")
        
        # Gem reference til destPath før vi lukker vinduet
        final_path = destPath
        
        # Deaktiver start-knappen permanent
        self.start_button.config(state='disabled')
        
        # Vis afslutningsbesked og spørg om at åbne mappe
        if messagebox.askyesno("Done", "Processing completed successfully!\nWould you like to open the output folder?"):
            try:
                os.startfile(final_path)
            except Exception as e:
                logging.error(f"Error opening output folder: {str(e)}", exc_info=True)
        
        # Luk vinduet til sidst
        try:
            self.root.quit()
            self.root.destroy()
        except Exception as e:
            logging.error(f"Error closing window: {str(e)}", exc_info=True)

def main():
    """Start ExcelCopyBOM programmet"""
    try:
        logging.basicConfig(filename='ExcelCopyBOM.log', level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
        app = QApplication(sys.argv)
        window = MainWindow()
        window.show()
        sys.exit(app.exec_())
    except Exception as e:
        logging.exception("Fatal error in main")
        print(f"Fatal error: {str(e)}")

if __name__ == "__main__":
    main()