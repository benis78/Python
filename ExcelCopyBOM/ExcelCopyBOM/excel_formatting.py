from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Farver
HEADER_FILL = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
ALTERNATE_ROW_FILL = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
ERROR_FILL = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

# Fonte
HEADER_FONT = Font(name='Calibri', size=11, bold=True, color="FFFFFF")
NORMAL_FONT = Font(name='Calibri', size=11)
BOLD_FONT = Font(name='Calibri', size=11, bold=True)

# Justering
CENTER_ALIGNMENT = Alignment(horizontal='center', vertical='center')
LEFT_ALIGNMENT = Alignment(horizontal='left', vertical='center')

# Kanter
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Kolonne bredder (i pixels)
# Note: Excel bruger en anden enhed for kolonnebredder, så vi konverterer fra pixels
# Cirka konvertering: pixels / 7.5 = Excel bredde enheder
COLUMN_WIDTHS = {
    'A': 52 / 7.5,  # Item
    'B': 152 / 7.5,  # Part Number
    'C': 47 / 7.5,  # Rev
    'D': 111 / 7.5,  # Description 1
    'E': 93 / 7.5,  # Thumbnail
    'F': 115 / 7.5,  # BOM Structure
    'G': 423 / 7.5,  # Description
    'H': 135 / 7.5,  # Material
    'I': 173 / 7.5,  # Title
    'J': 48 / 7.5,  # QTY
    'K': 82 / 7.5,  # Total QTY
    'L': 39 / 7.5,  # Diameter
    'M': 39 / 7.5,  # Thickness
    'N': 39 / 7.5,  # Length
    'O': 200 / 7.5,  # Keywords
    'P': 180 / 7.5,  # Type
    'Q': 125 / 7.5,  # Category
    'R': 94 / 7.5,  # Drawings
}

# Række højder (i pixels)
# Note: Excel bruger points for rækkehøjder, så vi konverterer fra pixels
# Cirka konvertering: pixels * 0.75 = points
HEADER_ROW_HEIGHT = 20 * 0.75  # 20 pixels for header
DATA_ROW_HEIGHT = 91 * 0.75    # 91 pixels for data rækker

# Kolonne justeringer
COLUMN_ALIGNMENTS = {
    'A': LEFT_ALIGNMENT,  # Item
    'B': LEFT_ALIGNMENT,    # Part Number
    'C': LEFT_ALIGNMENT,  # Rev
    'D': LEFT_ALIGNMENT,    # Description 1
    'E': LEFT_ALIGNMENT,    # Thumbnail
    'F': LEFT_ALIGNMENT,    # BOM Structure
    'G': LEFT_ALIGNMENT,    # Description
    'H': LEFT_ALIGNMENT,    # Material
    'I': LEFT_ALIGNMENT,    # Title
    'J': LEFT_ALIGNMENT,  # QTY
    'K': LEFT_ALIGNMENT,  # Total QTY
    'L': LEFT_ALIGNMENT,  # Diameter
    'M': LEFT_ALIGNMENT,  # Length
    'N': LEFT_ALIGNMENT,  # Thickness
    'O': LEFT_ALIGNMENT,    # Keywords
    'P': LEFT_ALIGNMENT,    # Type
    'Q': LEFT_ALIGNMENT,    # Category
    'R': LEFT_ALIGNMENT,    # Drawings
}

def apply_header_formatting(worksheet):
    """Anvender formatering på overskriftsrækken"""
    # Sæt header rækkehøjde
    worksheet.row_dimensions[1].height = HEADER_ROW_HEIGHT
    
    for cell in worksheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
        if cell.column_letter in COLUMN_ALIGNMENTS:
            cell.alignment = COLUMN_ALIGNMENTS[cell.column_letter]

def apply_data_formatting(worksheet):
    """Anvender formatering på datarækker"""
    for row_idx, row in enumerate(worksheet.iter_rows(min_row=2), start=2):
        # Sæt rækkehøjde for datarækker
        worksheet.row_dimensions[row_idx].height = DATA_ROW_HEIGHT
        
        # Skiftevis baggrundsfarve
        fill = ALTERNATE_ROW_FILL if row_idx % 2 == 0 else None
        
        for cell in row:
            cell.font = NORMAL_FONT
            cell.border = THIN_BORDER
            if fill:
                cell.fill = fill
            if cell.column_letter in COLUMN_ALIGNMENTS:
                cell.alignment = COLUMN_ALIGNMENTS[cell.column_letter]

def set_column_widths(worksheet):
    """Sætter kolonnebredder"""
    for column_letter, width in COLUMN_WIDTHS.items():
        worksheet.column_dimensions[column_letter].width = width

def format_worksheet(worksheet):
    """Anvender al formatering på et worksheet"""
    apply_header_formatting(worksheet)
    apply_data_formatting(worksheet)
    set_column_widths(worksheet) 