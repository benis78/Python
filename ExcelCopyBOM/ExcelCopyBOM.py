# import os, openpyxl, shutil, re, time
# import pandas as pd
# import tkinter as tk
# from tkinter import filedialog, messagebox
# import win32com.client as win32
# from concurrent.futures import ThreadPoolExecutor

import os
import shutil
import re
import time
import openpyxl
from tkinter import Tk, filedialog, messagebox
from win32com.client import Dispatch  # Importer kun Dispatch fra win32com.client
from concurrent.futures import ThreadPoolExecutor
from pandas import read_excel  # Importer kun read_excel fra pandas


root = Tk()
root.withdraw()

t='Open Excel BOM file'
file_path=filedialog.askopenfilename(title=t.upper(), filetypes=[('Excel files','.xlsx .xls')], initialdir='C:\\Working Folder\\Designs\\5-Projects')
#print(file_path)
#file_path='D:/Dropbox/Coding/BOM.xlsx'
efile=(os.path.basename(file_path))
ePath=(os.path.dirname(file_path))
eBOM=(os.path.abspath(file_path))
if file_path == '':
    os._exit(1)

# Tjek om kildefilen er skrivebeskyttet
if not os.access(file_path, os.W_OK):
    messagebox.showerror("ERROR", f"The Excel file '{efile}' is not writeable check the file out of the vault and try again.")
    os._exit(1)  # Stopper programmet her, hvis filen er skrivebeskyttet

# Find positionerne af de første 3 '-' i filnavnet
dash_positions = [pos for pos, char in enumerate(efile) if char == '-']

# Opret mappenavn ved at tage alt indtil det tredje '-' plus det første tegn efter det tredje '-'
if len(dash_positions) >= 3:
    # Find alt indtil det tredje '-'
    base_name = efile[:dash_positions[2] + 1]  # Inkluderer tredje '-'
    
    # Tag det fjerde tegn efter det tredje '-'
    if len(efile) > dash_positions[2] + 1:
        base_name += efile[dash_positions[2] + 1]  # Tilføj det første tegn efter tredje '-'

    new_folder_name = base_name
else:
    # Hvis der ikke er nok '-', brug hele filnavnet
    new_folder_name = efile.split('.')[0]  # Brug filnavnet uden udvidelse

# Definer destinationen for den nye mappe baseret på hvor Excel-filen ligger
dest_path = os.path.join(ePath, new_folder_name)

# Opret mappen, hvis den ikke allerede findes
if not os.path.exists(dest_path):
    os.makedirs(dest_path)

# #ask for destination path to drawings
# t='Save files in the Project\\3.0 Correspondence supplier\\Ipsen folder'
# dest_path=filedialog.askdirectory(title=t, initialdir='\\\\192.168.170.9\\Bigadan\\Bigadan\\1_PROJEKTER')
# #dest_path=''
destPath=(os.path.abspath(dest_path))
destDir=(os.path.basename)
if dest_path == '':
    os._exit(1)    
# #print(destPath)

#efile='2205-401-001-D - Digester Area.xlsx'
#destPath=('c:\\test\\drawings')
#eBOM=(r'c:\\test\\'+efile)
source = (r'\\192.168.170.18\drawings')
# source = (r'C:\Users\Jesper\Desktop\Ny mappe')
#eSource=('c:\\test')

start_time = time.time()

def newRev(name,files):
    #set the latestRev to the first object in the pdffiles list
    latestRev=files [0]
    #set the first revision character
    revChar='-'
    for rev in files:
        revPos = rev.find(str(name))+len(str(name))+1
        revLetter = rev[revPos]
        if revLetter > revChar:
            latestRev=rev
            revChar=revLetter
            # print (revChar)
    return latestRev, revChar


# Hjælpefunktion til at scanne individuelle entries i concurrent.futures
def scan_directory_task(entry):
    if entry.is_dir(follow_symlinks=False):
        # Hvis det er en mappe, så gå videre og scan undermapperne
        file_paths = []
        for sub_entry in os.scandir(entry.path):
            file_paths.extend(scan_directory_task(sub_entry))
        return file_paths
    elif entry.is_file(follow_symlinks=False):
        # Hvis det er en fil, sammenlign med targetFileNames og targetExtensions
        name = entry.name
        for targetName in targetFileNames:
            for targetExtension in targetExtensions:
                if name.startswith(targetName) and name.endswith(targetExtension):
                    if not name.endswith(excludeExtensions):
                        idx = -1
                        while True:
                            try:
                                idx = targetFileNames.index(targetName, idx + 1)
                                fileDestinations[idx].append(entry.path)
                            except ValueError:
                                break
        return [entry.path]
    return []

# Funktion til concurrent.futures scanning med os.scandir
def scan_directory_concurrent(directory):
    with ThreadPoolExecutor() as executor:
        futures = [executor.submit(scan_directory_task, entry) for entry in os.scandir(directory)]
        for future in futures:
            future.result()  # Vi skal vente på alle tråde at afslutte






# Opret forbindelse til Excel via COM
excel = Dispatch('Excel.Application')
excel.Visible = False  # Kør Excel i baggrunden
workbook = excel.Workbooks.Open(str(eBOM))
sheet = workbook.Sheets('BOM')

# Mønstre som skal fjernes i kolonne B
patterns_to_remove = ['0000-700', '0000-701', '0000-702', '0000-704', '0000-705']
pattern = re.compile(r'^(' + '|'.join(patterns_to_remove) + ')')

# Få det samlede antal brugte rækker i arket ved at bruge Excel's egen metode
max_row = sheet.UsedRange.Rows.Count

# Iterer gennem rækkerne og fjern dem, der matcher mønsteret eller starter med tekst i kolonne B
rows_to_delete = []
for row in range(2, max_row + 1):  # Start fra anden række for at undgå headers
    cell_value = str(sheet.Cells(row, 2).Value)  # Kolonne B

    # Fjern rækker, hvis cellen matcher mønstre eller starter med bogstav
    if pattern.match(cell_value) or (cell_value and cell_value[0].isalpha()):  # Tjek om første tegn er et bogstav
        rows_to_delete.append(row)

# Slet rækker (fra bunden op for at undgå skift i indeks)
for row in reversed(rows_to_delete):
    sheet.Rows(row).Delete()
efile_parts = efile.split('-')

if len(efile_parts) >= 4:  # Sørg for, at der er mindst 4 dele
    # Indsæt en ny række i Excel på række 2
    sheet.Rows(2).Insert()

    # Indsæt filnavnet (op til tredje '-') i B2
    sheet.Cells(2, 2).Value = '-'.join(efile_parts[:3])

    # Indsæt tallet 0 i A2
    sheet.Cells(2, 1).Value = 0

    # Indsæt bogstavet efter tredje '-' i C2, hvis det findes
    if efile_parts[3]:  # Tjek om der er et bogstav efter tredje '-'
        sheet.Cells(2, 3).Value = efile_parts[3][0]  # Første tegn efter tredje '-'
    else:
        sheet.Cells(2, 3).Value = "-"  # Hvis ingen bogstav findes, indsæt '-'

    # Indsæt tallet tekst Top assembly i G2
    sheet.Cells(2, 7).Value = "Top Assembly"

    # Indsæt tallet 1 i J2
    sheet.Cells(2, 10).Value = 1
# Gem og luk Excel-filen
workbook.Save()
workbook.Close(SaveChanges=True)
excel.Quit()

# os.chdir(ePath)
wb=openpyxl.load_workbook(eBOM)
sheet = wb['BOM']



# get max row count
max_row=sheet.max_row
#Start row
r=1      
# get max column count
max_column=sheet.max_column


targetFileNames=[]
for i in range(2,max_row+1):
  # iterate over all columns
  # get particular cell value
  cell_obj=sheet.cell(row=i,column=2)
  targetFileNames.append(cell_obj.value)
  #print (targetFileNames)


targetExtensions = ['.pdf']
excludeExtensions= '_FOR REVIEW.pdf'

fileDestinations = [[] for target in targetFileNames] # liste med samme længde som variablen targetFileNames, hvor hvert element til at starte med selv er en tom liste.
#print (fileDestinations)
# Værdi i variablen nu:
# fileDestinations = [[], [], []]

# for root, dirs, files in os.walk(source):           # Vi looper igennem vores directory
#     if 'Old' in dirs:
#         dirs.remove('Old')
#     if 'old' in dirs:
#         dirs.remove('old')
#     if 'OLD' in dirs:
#         dirs.remove('OLD')
#     for name in files:                              # Vi looper igennem hver fil i den aktuelle mappe
#         #Hver fil bliver så sammenlignet med hver værdi i targetFileNames med hver af de mulige extensions:
#         for targetName in targetFileNames:
#             for targetExtension in targetExtensions:
#                 if name.startswith(targetName) and name.endswith(targetExtension): 
#                     if not name.endswith(excludeExtensions):   
#                     # Når vi finder et match, bliver stien til den aktuelle fil tilføjet på den passende position i variablen fileDestinations:
#                         #fileDestinations[targetFileNames.index(targetName)]
#                         #fileDestinations[targetFileNames.index(targetName)].append(os.path.join(root,name))
#                         idx=-1
#                         while True:
#                             try:
#                                 idx = targetFileNames.index(targetName, idx+1)
#                                 fileDestinations[idx].append(os.path.join(root,name))
#                             except ValueError:
#                                 break
                       
                        #fileDestinations[ti].append(os.path.join(root,name))
                        #print(fileDestinations, end= '\n')
        
#print(targetName, end= '\n')
                    #print(name, end='\n') #Denne virker i for løkken og returere hele filnavnet men kun i løkken.


if __name__ == "__main__":
    # Start scanning med concurrent.futures
    scan_directory_concurrent(source)

#####################################

data=read_excel(file_path,sheet_name='BOM',header=0)

# Convert QTY To String
data['QTY'] = data['QTY'].astype(str)
# Convert string to int
#qty=[int(i) for i in data['QTY']]
qty = [int(ele) if ele.isdigit() else int(ele.rsplit(',', 1)[0]) for ele in data['QTY']] 



# Convert ITEM to String
data['Item']= data['Item'].astype(str)
# Remove .0 in Item column
data['Item']=[i.rsplit('.0', 1)[0] for i in data['Item']]
# Get parent level
parentlevels=[i.rsplit('.', 1)[0] for i in data['Item']]
# Get parent level with a empty string as first level I then can use in my for loop at the end.
pp=[i.rpartition('.')[0] for i in data['Item']]
#print(parentlevels)

# Put Items in a list called items
items=[]
for index, value in enumerate(data['Item']): 
    items.append(value)

# Search for items in  parentlevels 
ii=[items.index(i) for i in parentlevels]


totalQTY=[]
e=0
for i in pp:
    if i =='':
        totalQTY.append(qty[e])
        #print(i)
    else:
        totalQTY.append(qty[e]*totalQTY[ii[e]])
    e+=1

colQTY = data.columns.get_loc('QTY')

#tqty=data['Total QTY']=totalQTY

data.insert(colQTY+1,column='Total QTY',value=totalQTY)


#####################################



excelList=['' for target in targetFileNames]

revList=['' for target in targetFileNames]

i=0
for destination in fileDestinations:
    if len(destination) > 0:
        number,revChar = (newRev(targetFileNames[i],destination))
        excelList[i]=number
        revList[i]=revChar 
    i+=1


# set col to next
col=max_column+1

sheet.cell(row=1, column=col).value = 'Drawing'

# if not dest_path == '':
#     sheet.cell(row=1, column=col+1).value = 'Local Drawing'



i = 2
for File in excelList:
    if File == '':
        sheet.cell(row=i, column=col).value = 'Not available'
    else:
        shutil.copy2(File, destPath)  # Kopier filen til destinationsmappen
        localPath = os.path.basename(File)  # Brug kun filens navn, ikke stien

        # Indsæt hyperlink med kun filens navn
        sheet.cell(row=i, column=col).value = f'=HYPERLINK("{localPath}", "PDF")'
        sheet.cell(row=i, column=col).style = 'Hyperlink'
    
    i += 1


sheet.insert_cols(11,1)
sheet.cell(row=1, column=11).value = 'Total QTY'

i=2
for elements in totalQTY:
    sheet.cell(row=i, column=11).value = elements
    i+=1

        
    # Get column number of REV header
for col in sheet.iter_cols(1):
        for cell in col:
            if cell.value == "REV":
                c_rev=(sheet.cell(row=1, column=cell.column).column)
    
    # insert revion letter in excel
i=2
for File in revList:
    sheet.cell(row=i, column=c_rev).value = (File)
    i+=1




timestr = time.strftime("%Y%m%d-%H%M")


#timestr = time.strftime("%Y%m%d-%H%M")

 # copy drawings to ZIP folder
# if not dest_path=="":
    #  shutil.make_archive(destPath+'-'+'zip',destPath)
    # shutil.make_archive(destPath+'-'+'zip','zip')
# os.rename('a.txt', 'b.kml')

    # save excel         
wb.save(str(destPath)+'\\'+str(efile))

if not dest_path=="":
    shutil.make_archive(destPath+'-'+timestr,'zip',destPath)
    
#wb.save(efile)
# wb.save('D:/Dropbox/Coding/BOM_New.xlsx')
duration=(time.time() - start_time)

messagebox.showinfo(title='ExcelCopyBOM', message='Done\nTime: '+"{0:.2f}".format(duration)+' seconds')

# print("--- %s seconds ---" % (time.time() - start_time))

#Pyinstaller --onefile ExcelCopyBOM.py
#PyInstaller --onefile --upx-dir=D:\OneDrive\Coding\upx-4.2.4-win64 ExcelCopyBOM.py