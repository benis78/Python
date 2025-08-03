"""
Script til at generere en test BOM Excel fil
"""

import openpyxl
from openpyxl.styles import PatternFill, Font
import os

def create_test_bom():
    """Opretter en test BOM Excel fil"""
    wb = openpyxl.Workbook()
    ws = wb.active
    
    # Tilføj overskrifter
    headers = ['Item', 'Part Number', 'REV', 'BOM Structure', 'Description', 'QTY', 'D', 't', 'L']
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
        ws.cell(row=1, column=col).font = Font(bold=True)
        
    # Tilføj test data
    test_data = [
        # Normal komponenter
        ['1', '4003-02.1', 'A01', 'Inseparable', 'Main Assembly', '1', '100', '10', '1000'],
        ['1.1', '0000-301-123', '', 'Part', 'Basic Component 1', '2', '50', '5', '500'],
        ['1.2', '0000-302-456-A02', '', 'Part', 'Basic Component 2', '3', '30', '3', '300'],
        
        # Equipment rækker (skal fjernes hvis ikke inkluderet)
        ['2', '0000-701-789', '', 'Part', 'Equipment 1', '1', '200', '20', '2000'],
        ['2.1', '0000-702-012', '', 'Part', 'Equipment 2', '2', '150', '15', '1500'],
        
        # Phantom struktur
        ['3', '4003-03.2-B01', '', 'Phantom', 'Sub Assembly 1', '1', '80', '8', '800'],
        ['3.1', '0000-303-345', '', 'Part', 'Component 3', '4', '40', '4', '400'],
        ['3.2', '0000-304-678', '', 'Part', 'Component 4', '2', '60', '6', '600'],
        
        # Specielle rækker (kun bogstaver)
        ['4', 'ABCD', '', 'Part', 'Special Part', '1', '70', '7', '700'],
        
        # Inseparable med children (children skal fjernes)
        ['5', '0000-305-901', '', 'Inseparable', 'Inseparable Assembly', '1', '90', '9', '900'],
        ['5.1', '0000-306-234', '', 'Part', 'Child Component 1', '2', '45', '4.5', '450'],
        ['5.2', '0000-307-567', '', 'Part', 'Child Component 2', '3', '35', '3.5', '350'],
        
        # Basic Component med children (skal fjernes)
        ['6', '0000-308-890', '', 'Part', 'Basic Component Assembly', '1', '110', '11', '1100'],
        ['6.1', '0000-309-123', '', 'Part', 'Child Basic 1', '2', '55', '5.5', '550'],
        ['6.2', '0000-310-456', '', 'Part', 'Child Basic 2', '3', '65', '6.5', '650']
    ]
    
    # Indsæt test data
    for row_idx, row_data in enumerate(test_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)
            
    # Gem filen
    filename = os.path.join(os.path.dirname(__file__), '4003-02.1-A01-- - BOM.xlsx')
    wb.save(filename)
    print(f"Test BOM fil oprettet: {filename}")
    
if __name__ == "__main__":
    create_test_bom() 