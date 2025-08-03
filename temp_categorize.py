import pandas as pd
from Categories import CategoryParser
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.drawing.image import Image
import os
import shutil
import win32com.client as win32
import glob

def find_file_in_subdirs(base_dir, filename):
    """Søg efter fil i alle undermapper"""
    for root, dirs, files in os.walk(base_dir):
        if filename in files:
            return os.path.join(root, filename)
    return None

def find_drawing_files(base_dir, part_number):
    """Find alle relevante tegningsfiler for et part number"""
    pattern = f"{part_number}*.pdf"
    matches = []
    for root, dirs, files in os.walk(base_dir):
        for file in files:
            if file.startswith(part_number) and file.endswith('.pdf'):
                matches.append(os.path.join(root, file))
    return matches

def copy_with_hyperlinks(source_file, output_file):
    # Brug Excel COM object til at kopiere med hyperlinks
    excel = win32.Dispatch('Excel.Application')
    excel.Visible = False
    wb = excel.Workbooks.Open(source_file)
    wb.SaveAs(output_file)
    wb.Close()
    excel.Quit()

# Læs original BOM-fil med hyperlinks
source_file = os.path.abspath('C:/Coding/Python/Cursor_ExcelCopyBOM/4003-02.1-A01-- - BOM.xlsx')
output_file = os.path.abspath('C:/Coding/Python/Cursor_ExcelCopyBOM/4003-02.1-A01-- - BOM_Categorized.xlsx')
source_dir = 'C:/Coding/Python/ExcelCopyBOM/Files'

# Kopier original fil med hyperlinks
copy_with_hyperlinks(source_file, output_file)

# Åbn den kopierede fil
wb = openpyxl.load_workbook(output_file)
source_sheet = wb['BOM']

# Læs data med pandas for kategorisering
df = pd.read_excel(source_file, sheet_name='BOM')

# Initialiser CategoryParser
parser = CategoryParser()

# Gruppér data efter kategori
categorized_data = {}
for idx, row in df.iterrows():
    if pd.notna(row['Part Number']):
        category, type_ = parser.categorize(str(row['Part Number']))
        if category not in categorized_data:
            categorized_data[category] = {'rows': [], 'indices': []}
        categorized_data[category]['rows'].append(row)
        categorized_data[category]['indices'].append(idx + 2)  # +2 for Excel's 1-baserede indeks og header

# Opret nye faner og kopier data med hyperlinks
for category, data in categorized_data.items():
    # Opret ny fane
    if category not in wb.sheetnames:
        wb.create_sheet(category)
    category_sheet = wb[category]
    
    # Kopier header
    for col in range(1, source_sheet.max_column + 1):
        category_sheet.cell(1, col, source_sheet.cell(1, col).value)
    
    # Kopier rækker med hyperlinks og formatering
    for new_row, (excel_row, row_data) in enumerate(zip(data['indices'], data['rows']), start=2):
        for col in range(1, source_sheet.max_column + 1):
            source_cell = source_sheet.cell(excel_row, col)
            target_cell = category_sheet.cell(new_row, col)
            
            # Kopier værdi
            target_cell.value = source_cell.value
            
            # Kopier hyperlink hvis det findes
            if source_cell.hyperlink:
                target_cell.hyperlink = source_cell.hyperlink
                target_cell.style = source_cell.style
    
    # Tilpas kolonnebredder
    for column in category_sheet.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        category_sheet.column_dimensions[column[0].column_letter].width = adjusted_width

# Slet den originale BOM fane
if 'BOM' in wb.sheetnames:
    del wb['BOM']

# Gem Excel-filen
wb.save(output_file)

# Opret mapper for hver kategori og kopier tegninger
base_dir = 'C:/Coding/Python/Cursor_ExcelCopyBOM/Categorized_Drawings'
os.makedirs(base_dir, exist_ok=True)

# Kopier tegninger til deres respektive kategorimapper
for category, data in categorized_data.items():
    category_dir = os.path.join(base_dir, category)
    os.makedirs(category_dir, exist_ok=True)
    
    # Find alle tegninger for denne kategori
    for row in data['rows']:
        if pd.notna(row['Part Number']):
            try:
                # Søg efter tegninger baseret på part number
                drawing_files = find_drawing_files(source_dir, str(row['Part Number']))
                
                for src_path in drawing_files:
                    if os.path.exists(src_path):
                        dest_path = os.path.join(category_dir, os.path.basename(src_path))
                        shutil.copy2(src_path, dest_path)
                        print(f"Kopieret tegning: {os.path.basename(src_path)} til {category}")
                
                if not drawing_files:
                    print(f"Ingen tegninger fundet for: {row['Part Number']}")
            except Exception as e:
                print(f"Fejl ved kopiering af tegninger for {row['Part Number']} til {category_dir}: {str(e)}")

print(f"Kategoriseret BOM er gemt som: {output_file}")
print(f"Tegninger er organiseret i: {base_dir}") 