import openpyxl
import re
import os
from pathlib import Path
import tkinter as tk
from tkinter import filedialog, messagebox, ttk

class DescriptionSplitter:
    def __init__(self):
        self.root = tk.Tk()
        self.root.title("Description Splitter")
        
        # Sæt vinduesstørrelse
        window_width = 500
        window_height = 250  # Øget højde for at få plads til sheet vælger
        
        # Beregn position for at centrere vinduet
        screen_width = self.root.winfo_screenwidth()
        screen_height = self.root.winfo_screenheight()
        x = (screen_width - window_width) // 2
        y = (screen_height - window_height) // 2
        
        # Sæt vinduesstørrelse og position
        self.root.geometry(f'{window_width}x{window_height}+{x}+{y}')
        
        # GUI elementer
        frame = ttk.Frame(self.root, padding="10")
        frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # Excel fil vælger
        ttk.Label(frame, text="Excel File:").grid(row=0, column=0, sticky=tk.W)
        self.file_path_var = tk.StringVar()
        self.file_entry = ttk.Entry(frame, textvariable=self.file_path_var, width=50)
        self.file_entry.grid(row=0, column=1, padx=5)
        ttk.Button(frame, text="Browse", command=self.browse_file).grid(row=0, column=2)
        
        # Procesbar
        self.progress_var = tk.DoubleVar()
        self.progress = ttk.Progressbar(frame, length=400, mode='determinate', variable=self.progress_var)
        self.progress.grid(row=1, column=0, columnspan=3, pady=20)
        
        # Status label
        self.status_var = tk.StringVar(value="Ready to start...")
        self.status_label = ttk.Label(frame, textvariable=self.status_var)
        self.status_label.grid(row=2, column=0, columnspan=3)
        
        # Sheet vælger
        ttk.Label(frame, text="Select Sheets:").grid(row=3, column=0, sticky=tk.W, pady=(10,0))
        self.sheet_var = tk.StringVar(value="All Sheets")
        self.sheet_combo = ttk.Combobox(frame, textvariable=self.sheet_var, state="readonly", width=30)
        self.sheet_combo.grid(row=3, column=1, padx=5, pady=(10,0))
        
        # Start knap
        self.start_button = ttk.Button(frame, text="Start Processing", command=self.start_processing)
        self.start_button.grid(row=4, column=0, columnspan=3, pady=10)

    def browse_file(self):
        file_path = filedialog.askopenfilename(
            title="OPEN EXCEL FILE",
            filetypes=[('Excel files', '.xlsx .xls')]
        )
        if file_path:
            self.file_path_var.set(file_path)
            self.update_sheet_list(file_path)

    def update_progress(self, value, message):
        self.progress_var.set(value)
        self.status_var.set(message)
        self.root.update()

    def find_size_patterns(self, text):
        """Finder alle størrelsesbetegnelser i teksten"""
        patterns = {
            'DN': r'DN(\d+)',
            'Ø': r'[Øø](\d+(?:,\d+)?)',
            'SDR': r'SDR(\d+(?:,\d+)?)',
            'PN': r'PN(\d+)',
            'DVR': r'DVR(\d+(?:,\d+)?)'
        }
        
        found_sizes = {}
        for size_type, pattern in patterns.items():
            matches = re.findall(pattern, text)
            if matches:
                # Tag den første match for hver type
                found_sizes[size_type] = matches[0]
        
        return found_sizes

    def split_description(self, description):
        """Splitter beskrivelsen i grupper baseret på ' - ' separator"""
        if not description:
            return []
        
        # Split på ' - ' (mellemrum bindestreg mellemrum)
        groups = description.split(' - ')
        return [group.strip() for group in groups if group.strip()]

    def find_column_by_name(self, sheet, column_name):
        """Finder kolonnenummer for en given kolonnenavn"""
        for col in range(1, sheet.max_column + 1):
            if sheet.cell(row=1, column=col).value == column_name:
                return col
        return None

    def get_next_empty_column(self, sheet):
        """Finder den næste tomme kolonne efter den sidste brugte kolonne"""
        return sheet.max_column + 1

    def update_sheet_list(self, file_path):
        """Opdaterer dropdown listen med sheets fra den valgte fil"""
        try:
            wb = openpyxl.load_workbook(file_path, read_only=True)
            sheets = wb.sheetnames
            wb.close()
            
            # Tilføj "All Sheets" som første mulighed
            sheet_options = ["All Sheets"] + sheets
            self.sheet_combo['values'] = sheet_options
            self.sheet_var.set("All Sheets")
            
        except Exception as e:
            messagebox.showerror("Error", f"Could not read Excel file: {str(e)}")
            self.sheet_combo['values'] = ["All Sheets"]
            self.sheet_var.set("All Sheets")

    def process_excel_file(self, file_path):
        """Hovedfunktionen der behandler Excel-filen"""
        try:
            # Åbn Excel-filen
            wb = openpyxl.load_workbook(file_path)
            
            # Bestem hvilke sheets der skal behandles
            selected_sheet = self.sheet_var.get()
            if selected_sheet == "All Sheets":
                sheets_to_process = wb.sheetnames
            else:
                sheets_to_process = [selected_sheet]
            
            # Behandl de valgte sheets
            for sheet_name in sheets_to_process:
                sheet = wb[sheet_name]
                self.update_progress(10, f"Processing sheet: {sheet_name}")
                
                # Find kolonnerne
                description_col = self.find_column_by_name(sheet, 'Description')
                part_number_col = self.find_column_by_name(sheet, 'Part Number')
                
                if not description_col or not part_number_col:
                    print(f"Warning: Required columns not found in sheet {sheet_name}")
                    continue
                
                # Find kolonne med header "D" (for Ø værdier)
                d_col = self.find_column_by_name(sheet, 'D')
                
                # Find den sidste brugte kolonne
                last_used_col = sheet.max_column
                
                # Opret nye kolonner hvis nødvendigt
                designation_col = last_used_col + 1
                sheet.cell(row=1, column=designation_col).value = 'Designation'
                
                # Formater header-rækken for Designation kolonnen
                designation_cell = sheet.cell(row=1, column=designation_col)
                designation_cell.font = openpyxl.styles.Font(bold=True)
                designation_cell.fill = openpyxl.styles.PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
                
                # Tilføj auto-filter til Designation kolonnen
                sheet.auto_filter.ref = f"A1:{sheet.cell(row=1, column=designation_col).column_letter}1"
                
                # Dictionary til at holde styr på nye kolonner for størrelser
                size_columns = {}
                
                # Gennemgå alle rækker (start fra række 2, da række 1 er header)
                for row in range(2, sheet.max_row + 1):
                    part_number = str(sheet.cell(row=row, column=part_number_col).value or '')
                    description = str(sheet.cell(row=row, column=description_col).value or '')
                    
                    # Tjek om Part Number starter med 0000-7
                    if not part_number.startswith('0000-7'):
                        continue
                    
                    # Split beskrivelsen
                    groups = self.split_description(description)
                    if not groups:
                        continue
                    
                    # Første gruppe er Designation
                    if groups:
                        sheet.cell(row=row, column=designation_col).value = groups[0]
                    
                    # Find størrelsesbetegnelser i hele beskrivelsen
                    size_patterns = self.find_size_patterns(description)
                    
                    # Behandl hver størrelsesbetegnelse
                    for size_type, size_value in size_patterns.items():
                        if size_type == 'Ø':
                            # Ø værdier går i kolonne med header "D" (kun tallet, ikke Ø tegnet)
                            if d_col:
                                sheet.cell(row=row, column=d_col).value = size_value
                        else:
                            # Andre størrelser får deres egen kolonne
                            if size_type not in size_columns:
                                # Opret ny kolonne for denne størrelsesbetegnelse
                                new_col = self.get_next_empty_column(sheet)
                                sheet.cell(row=1, column=new_col).value = size_type
                                
                                # Formater header-rækken for den nye kolonne
                                header_cell = sheet.cell(row=1, column=new_col)
                                header_cell.font = openpyxl.styles.Font(bold=True)
                                header_cell.fill = openpyxl.styles.PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
                                
                                # Opdater auto-filter til at inkludere den nye kolonne
                                sheet.auto_filter.ref = f"A1:{sheet.cell(row=1, column=new_col).column_letter}1"
                                
                                size_columns[size_type] = new_col
                            
                            # Indsæt værdien
                            sheet.cell(row=row, column=size_columns[size_type]).value = f"{size_type}{size_value}"
                
                # Juster kolonnebredder
                for col in sheet.columns:
                    max_length = 0
                    column = col[0].column_letter
                    for cell in col:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = (max_length + 2)
                    sheet.column_dimensions[column].width = adjusted_width
                
                self.update_progress(50, f"Completed sheet: {sheet_name}")
            
            # Gem filen med nyt navn
            file_path_obj = Path(file_path)
            new_filename = f"{file_path_obj.stem}_processed{file_path_obj.suffix}"
            new_file_path = file_path_obj.parent / new_filename
            
            wb.save(new_file_path)
            self.update_progress(100, "Processing completed!")
            
            return str(new_file_path)
            
        except Exception as e:
            raise Exception(f"Error processing Excel file: {str(e)}")

    def start_processing(self):
        file_path = self.file_path_var.get()
        
        if not file_path:
            messagebox.showerror("Error", "Please select an Excel file first.")
            return
            
        if not os.path.exists(file_path):
            messagebox.showerror("Error", "The selected file does not exist.")
            return
            
        self.start_button.config(state='disabled')
        self.update_progress(0, "Starting process...")
        
        try:
            output_file = self.process_excel_file(file_path)
            messagebox.showinfo("Success", f"Processing completed!\nOutput saved as:\n{output_file}")
        except Exception as e:
            messagebox.showerror("Error", f"An error occurred: {str(e)}")
        finally:
            self.start_button.config(state='normal')

def main():
    app = DescriptionSplitter()
    app.root.mainloop()

if __name__ == "__main__":
    main() 