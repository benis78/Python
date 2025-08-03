# import os, openpyxl, shutil, re, time
# import pandas as pd
# import tkinter as tk
# from tkinter import filedialog, messagebox
# import win32com.client as win32
# from concurrent.futures import ThreadPoolExecutor

import os
import shutil
import re
import time
import openpyxl
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from win32com.client import Dispatch  # Importer kun Dispatch fra win32com.client
from concurrent.futures import ThreadPoolExecutor
from pandas import read_excel  # Importer kun read_excel fra pandas
import logging  # Tilføj logging
#from ECB_grouping import group_by_parent_items
import win32com.client
import pythoncom
from typing import Dict, List, Optional, Tuple
import tempfile
from pathlib import Path
import pandas as pd
from Categories import CategoryParser
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.drawing.image import Image
import glob
import sqlite3
from datetime import datetime

# Globale variabler
targetFileNames = []
targetExtensions = []
excludeExtensions = ''
fileDestinations = []
root_files = {}  # Dictionary til at holde styr på filer der skal i rodmappen

# Konfigurer logging med både fil og console output
log_file = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'ExcelCopyBOM.log')

# Opret en logger
logger = logging.getLogger('ExcelCopyBOM')
logger.setLevel(logging.DEBUG)

# Fjern eksisterende handlers hvis nogen
for handler in logger.handlers[:]:
    logger.removeHandler(handler)

# Opret en formatter
formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')

# Opret en fil handler
file_handler = logging.FileHandler(log_file, mode='w')
file_handler.setLevel(logging.DEBUG)
file_handler.setFormatter(formatter)

# Tilføj fil handler til logger
logger.addHandler(file_handler)

# Test logging
logger.info(f"Log file location: {log_file}")
logger.info("Starting ExcelCopyBOM program")
logger.debug("Debug test message")
logger.info("Info test message")
logger.warning("Warning test message")
logger.error("Error test message")

# Erstat alle logging kald med logger
logging = logger

def find_drawing_files(directory, part_number):
    """Find alle PDF filer der starter med det givne part number i den angivne mappe og undermapper"""
    pdf_files = []
    try:
        # Tjek om netværksmappen er tilgængelig
        if not os.path.exists(directory):
            logger.error(f"Netværksmappen {directory} er ikke tilgængelig")
            return pdf_files
            
        # Tjek om vi har læseadgang til mappen
        if not os.access(directory, os.R_OK):
            logger.error(f"Ingen læseadgang til mappen {directory}")
            return pdf_files
            
        logger.info(f"Søger efter tegninger for {part_number} i {directory}")
        
        for root, dirs, files in os.walk(directory):
            # Ignorer mapper med 'Old' i navnet
            if 'Old' in dirs:
                dirs.remove('Old')
            if 'old' in dirs:
                dirs.remove('old')
            if 'OLD' in dirs:
                dirs.remove('OLD')
                
            for file in files:
                if file.startswith(part_number) and file.endswith('.pdf') and not file.endswith('_FOR REVIEW.pdf'):
                    full_path = os.path.join(root, file)
                    pdf_files.append(full_path)
                    logger.info(f"Fundet tegning: {full_path}")
                    
        if not pdf_files:
            logger.warning(f"Ingen tegninger fundet for {part_number} i {directory}")
            
    except Exception as e:
        logger.error(f"Fejl ved søgning efter tegninger: {str(e)}")
        
    return pdf_files

class ExcelCopyBOMGUI:
    def __init__(self):
        self.root = tk.Tk()
        self.root.title("Excel Copy BOM")
        # Fjern topmost attribut
        # self.root.attributes('-topmost', True)
        
        # Sæt vinduesstørrelse
        window_width = 500
        window_height = 350  # Øget højde for at få plads til nye elementer
        
        # Beregn position for at centrere vinduet
        screen_width = self.root.winfo_screenwidth()
        screen_height = self.root.winfo_screenheight()
        x = (screen_width - window_width) // 2
        y = (screen_height - window_height) // 2
        
        # Sæt vinduesstørrelse og position
        self.root.geometry(f'{window_width}x{window_height}+{x}+{y}')
        
        # Excel fil vælger
        frame = ttk.Frame(self.root, padding="10")
        frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # Excel BOM File section
        ttk.Label(frame, text="Excel BOM File:").grid(row=0, column=0, sticky=tk.W)
        self.file_path_var = tk.StringVar()
        self.file_entry = ttk.Entry(frame, textvariable=self.file_path_var, width=50)
        self.file_entry.grid(row=0, column=1, padx=5)
        ttk.Button(frame, text="Browse", command=self.browse_file).grid(row=0, column=2)
        
        # Source Folder section
        ttk.Button(frame, text="Source Folder", command=lambda: self.source_path_var.set(r'\\192.168.170.18\drawings')).grid(row=1, column=0, sticky=tk.W, pady=(10,0))
        self.source_path_var = tk.StringVar(value=r'\\192.168.170.18\drawings')
        self.source_entry = ttk.Entry(frame, textvariable=self.source_path_var, width=50)
        self.source_entry.grid(row=1, column=1, padx=5, pady=(10,0))
        ttk.Button(frame, text="Browse", command=self.browse_source).grid(row=1, column=2, pady=(10,0))
        
        # Search Method section
        search_frame = ttk.LabelFrame(frame, text="Search Method", padding="5")
        search_frame.grid(row=2, column=0, columnspan=3, pady=10, sticky=(tk.W, tk.E))
        
        self.search_method = tk.StringVar(value="database")  # Sæt database som default
        ttk.Radiobutton(search_frame, text="Database Search", variable=self.search_method, 
                       value="database").grid(row=0, column=0, padx=20)
        ttk.Radiobutton(search_frame, text="Direct Search", variable=self.search_method, 
                       value="direct").grid(row=0, column=1, padx=20)
        
        # Database Last Updated
        self.db_status = tk.StringVar(value="Database last updated: Checking...")
        ttk.Label(search_frame, textvariable=self.db_status).grid(row=1, column=0, columnspan=2, pady=5)
        
        # Tilføj refresh knap
        ttk.Button(search_frame, text="Refresh", command=self.check_database_status).grid(row=1, column=2, padx=5)
        
        # Update database status
        self.check_database_status()
        
        # DWG Checkbox
        self.include_dwg = tk.BooleanVar(value=True)
        ttk.Checkbutton(frame, text="Include DWG files", variable=self.include_dwg).grid(row=3, column=0, columnspan=3, pady=5)
        
        # Procesbar
        self.progress_var = tk.DoubleVar()
        self.progress = ttk.Progressbar(frame, length=400, mode='determinate', variable=self.progress_var)
        self.progress.grid(row=4, column=0, columnspan=3, pady=20)
        
        # Status label
        self.status_var = tk.StringVar(value="Ready to start...")
        self.status_label = ttk.Label(frame, textvariable=self.status_var)
        self.status_label.grid(row=5, column=0, columnspan=3)
        
        # Start knap
        self.start_button = ttk.Button(frame, text="Start Processing", command=self.start_processing)
        self.start_button.grid(row=6, column=0, columnspan=3, pady=10)

    def browse_source(self):
        folder_path = filedialog.askdirectory(
            title="SELECT SOURCE FOLDER",
            initialdir=self.source_path_var.get()
        )
        if folder_path:
            self.source_path_var.set(folder_path)

    def browse_file(self):
        file_path = filedialog.askopenfilename(
            title="OPEN EXCEL BOM FILE",
            filetypes=[('Excel files', '.xlsx .xls')],
            initialdir='C:\\Working Folder\\Designs\\5-Projects'
        )
        if file_path:
            self.file_path_var.set(file_path)

    def newRev(self, name, files):
            # Opdel filer efter filtype
            pdf_files = [f for f in files if f.endswith('.pdf')]
            dwg_files = [f for f in files if f.endswith('.dwg')]
            
            # Initialiser returværdier
            latestFiles = []
            revChar = '-'
            
            # Find seneste PDF
            if pdf_files:
                latestPdf = pdf_files[0]
                for rev in pdf_files:
                    revPos = rev.find(str(name)) + len(str(name)) + 1
                    revLetter = rev[revPos]
                    if revLetter > revChar:
                        latestPdf = rev
                        revChar = revLetter
                latestFiles.append(latestPdf)
            
            # Find seneste DWG med samme revision hvis DWG er aktiveret
            if self.include_dwg.get() and dwg_files:
                # Find DWG med samme revision som PDF
                matching_dwg = None
                for dwg in dwg_files:
                    revPos = dwg.find(str(name)) + len(str(name)) + 1
                    if dwg[revPos] == revChar:
                        matching_dwg = dwg
                        break
                
                # Hvis ingen DWG med samme revision findes, tag den seneste DWG
                if not matching_dwg:
                    matching_dwg = dwg_files[0]
                    for rev in dwg_files:
                        revPos = rev.find(str(name)) + len(str(name)) + 1
                        revLetter = rev[revPos]
                        if revLetter > revChar:
                            matching_dwg = rev
                            revChar = revLetter
                
                if matching_dwg:
                    latestFiles.append(matching_dwg)
            
            return latestFiles, revChar

    def update_progress(self, value, message):
        self.progress_var.set(value)
        self.status_var.set(message)
        self.root.update()

    def start_processing(self):
        file_path = self.file_path_var.get()
        source_path = self.source_path_var.get()
        
        if not file_path:
            messagebox.showerror("Error", "Please select an Excel BOM file first.")
            return
            
        if not source_path:
            messagebox.showerror("Error", "Please select a Source folder first.")
            return
            
        if not os.path.exists(source_path):
            messagebox.showerror("Error", "The selected Source folder does not exist.")
            return
            
        self.start_button.config(state='disabled')
        self.update_progress(0, "Starting process...")
        
        try:
            self.process_bom(file_path, source_path)
        except Exception as e:
            logger.error(f"Error during processing: {str(e)}", exc_info=True)
            messagebox.showerror("Error", f"An error occurred: {str(e)}")
            self.start_button.config(state='normal')

    def should_include_row(self, row, df):
        """Afgør om en række skal inkluderes baseret på BOM Structure regler."""
        item_number = str(row['Item'])
        bom_structure = str(row.get('BOM Structure', '')).strip().upper()
        
        # Hvis rækken er markeret som Phantom, skal den ikke inkluderes
        if bom_structure == "PHANTOM":
            return False
            
        # Hvis denne række er Inseparable, skal den inkluderes
        if bom_structure == "INSEPARABLE":
            return True
            
        # Hvis rækken har en parent, tjek ALLE parents i hierarkiet
        if '.' in item_number:
            parts = item_number.split('.')
            # Start med at tjekke den nærmeste parent og gå op gennem hierarkiet
            for i in range(len(parts)-1, 0, -1):
                parent_number = '.'.join(parts[:i])
                parent_mask = df['Item'] == parent_number
                if parent_mask.any():
                    parent_row = df[parent_mask].iloc[0]
                    parent_structure = str(parent_row.get('BOM Structure', '')).strip().upper()
                    # Hvis NOGEN parent er Inseparable, skal denne række ikke inkluderes
                    if parent_structure == "INSEPARABLE":
                        return False
        
        return True

    def remove_empty_folders(self, path):
        """Fjerner alle tomme mapper i den givne sti."""
        for root, dirs, files in os.walk(path, topdown=False):
            for dirname in dirs:
                dir_path = os.path.join(root, dirname)
                try:
                    # Tjek om mappen er tom (ingen filer og ingen undermapper)
                    if not os.listdir(dir_path):
                        os.rmdir(dir_path)
                        logger.info(f"Fjernede tom mappe: {dir_path}")
                except Exception as e:
                    logger.error(f"Kunne ikke fjerne tom mappe {dir_path}: {str(e)}")

    def adjust_column_widths(self, sheet):
        """Juster kolonnebredder og formater header-rækken for en given sheet"""
        # Formater header-rækken
        for cell in sheet[1]:
            cell.font = openpyxl.styles.Font(bold=True)
            cell.fill = openpyxl.styles.PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        
        # Tilføj filter
        sheet.auto_filter.ref = sheet.dimensions
        
        # Fryse den første række
        sheet.freeze_panes = 'A2'
        
        # Juster kolonnebredder til sidst
        for col in sheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 3.5)  # Tilføjet 1.5 ekstra til bredden
            sheet.column_dimensions[column].width = adjusted_width

    def group_rows_by_level(self, sheet):
        """Grupperer rækker baseret på niveau i Item kolonnen (baseret på antal punktummer)"""
        # Find Item kolonnen
        item_col = None
        for col in range(1, sheet.max_column + 1):
            if sheet.cell(row=1, column=col).value == "Item":
                item_col = col
                break
        
        if not item_col:
            return
        
        # Opsaml alle niveauer og deres rækker
        levels = {}
        for row in range(2, sheet.max_row + 1):
            item = str(sheet.cell(row=row, column=item_col).value)
            if not item:
                continue
                
            # Tæl antal punktummer for at bestemme niveau
            level = item.count('.')
            if level not in levels:
                levels[level] = []
            levels[level].append(row)
        
        # Grupper rækker fra højeste til laveste niveau
        for level in sorted(levels.keys(), reverse=True):
            if level == 0:  # Skip top-niveau
                continue
                
            for row in levels[level]:
                sheet.row_dimensions.group(row, row, outline_level=level)

    def scan_directory_concurrent(self, directory):
        """Concurrent scanning af mappe med ThreadPoolExecutor"""
        with ThreadPoolExecutor() as executor:
            futures = [executor.submit(self.scan_directory_task, entry) for entry in os.scandir(directory)]
            for future in futures:
                future.result()  # Vi skal vente på alle tråde at afslutte

    def process_bom(self, file_path, source_path):
        global targetFileNames, targetExtensions, excludeExtensions, fileDestinations
        
        try:
            # Eksisterende variabler
            efile = os.path.basename(file_path)
            ePath = os.path.dirname(file_path)
            eBOM = os.path.abspath(file_path)
            
            self.update_progress(20, "Creating destination folder...")
            
            # Find positionerne af de første 3 '-' i filnavnet
            dash_positions = [pos for pos, char in enumerate(efile) if char == '-']
            
            # Opret mappenavn ved at tage alt indtil det tredje '-' plus tegnet efter det tredje '-'
            if len(dash_positions) >= 3:
                base_name = efile[:dash_positions[2]]
                if len(efile) > dash_positions[2] + 1:
                    base_name += efile[dash_positions[2]:dash_positions[2] + 2]
                new_folder_name = base_name.strip()
            else:
                new_folder_name = os.path.splitext(efile)[0]

            # Definer destinationen for den nye mappe
            dest_path = os.path.join(ePath, new_folder_name)
            os.makedirs(dest_path, exist_ok=True)
            
            destPath = os.path.abspath(dest_path)
            destDir = os.path.basename(dest_path)
            
            if dest_path == '':
                raise ValueError("Destination path is empty")

            source = source_path
            start_time = time.time()

            # Initialiser targetFileNames
            targetFileNames.clear()
            wb = openpyxl.load_workbook(file_path)
            source_sheet = wb['BOM']
            max_row = source_sheet.max_row
            
            for i in range(2, max_row + 1):
                cell_obj = source_sheet.cell(row=i, column=2)
                if cell_obj.value:
                    targetFileNames.append(str(cell_obj.value))

            # Ekstraher det forkortede filnavn (minus 6 tegn) fra input filen
            base_filename = os.path.splitext(efile)[0]  # Fjern .xlsx
            root_file_prefix = base_filename[:-6]  # Tag alt undtagen de sidste 6 tegn
            if root_file_prefix:
                # Tilføj det som et særligt element i targetFileNames
                targetFileNames.append(root_file_prefix)
                # Opret en global variabel til at holde styr på hvilke filer der skal i rodmappen
                global root_files
                root_files = {root_file_prefix: True}  # Dictionary til at markere root files
                logger.info(f"Added root file prefix to search: {root_file_prefix}")

            # Definer filtyper
            targetExtensions.clear()
            targetExtensions.append('.pdf')
            if self.include_dwg.get():
                targetExtensions.append('.dwg')
            
            excludeExtensions = '_FOR REVIEW.pdf'

            # Start søgning baseret på valgt metode
            self.update_progress(30, "Searching for files...")
            search_success = False
            
            if self.search_method.get() == "database":
                self.status_var.set("Searching in database...")
                search_success = self.search_files_in_database(targetFileNames)
            else:
                self.status_var.set("Searching directly in folders...")
                search_success = self.search_files_directly(source)

            if not search_success:
                raise Exception("File search failed")

            self.update_progress(50, "Processing files...")
            
            # Læs data med pandas for kategorisering
            df = pd.read_excel(file_path, sheet_name='BOM')

            # Convert QTY To String
            df['QTY'] = df['QTY'].astype(str)
            # Convert string to int
            qty = [int(ele) if ele.isdigit() else int(ele.rsplit(',', 1)[0]) for ele in df['QTY']] 

            # Fjern .0 og trim whitespace fra Item kolonnen
            df['Item'] = df['Item'].astype(str)
            df['Item'] = [i.rsplit('.0', 1)[0].strip() for i in df['Item']]

            # Generer parentlevels
            parentlevels = [i.rsplit('.', 1)[0].strip() for i in df['Item']]

            # Opret listen over items
            items = [value.strip() for value in df['Item']]

            # Find index for parentlevels i items
            ii = []
            for i in parentlevels:
                if i == '':
                    ii.append(None)  # Ingen parent for top-level items
                else:
                    try:
                        ii.append(items.index(i))
                    except ValueError:
                        logger.error(f"Parent item '{i}' not found in items list")
                        ii.append(None)  # Hvis parent ikke findes, tilføj None

            # Beregn Total QTY
            totalQTY = []
            e = 0
            for i in parentlevels:
                if i == '':
                    totalQTY.append(qty[e])
                else:
                    if ii[e] is not None:
                        try:
                            totalQTY.append(qty[e] * totalQTY[ii[e]])
                        except IndexError:
                            logger.error(f"IndexError: Unable to calculate Total QTY for row {e + 2}. Parent index {ii[e]} is out of range.")
                            totalQTY.append(qty[e])  # Brug kun egen QTY hvis der er en fejl
                    else:
                        logger.warning(f"Parent not found for row {e + 2}. Using only own QTY.")
                        totalQTY.append(qty[e])  # Hvis parent ikke findes, brug kun egen QTY
                e += 1

            # Indsæt Total QTY kolonne i BOM fanen
            source_sheet.insert_cols(11, 1)
            source_sheet.cell(row=1, column=11).value = 'Total QTY'
            
            # Indsæt Total QTY værdier i BOM fanen
            for i, qty in enumerate(totalQTY, start=2):
                source_sheet.cell(row=i, column=11).value = qty

            # Find REV kolonnens position
            c_rev = None
            for col in source_sheet.iter_cols(1):
                for cell in col:
                    if cell.value == "REV":
                        c_rev = cell.column
                        break
                if c_rev:
                    break

            # Beregn revisionsbogstaver
            excelList = [[] for _ in targetFileNames]
            revList = ['' for _ in targetFileNames]

            for i, destination in enumerate(fileDestinations):
                if destination:
                    files, revChar = self.newRev(targetFileNames[i], destination)
                    excelList[i] = files
                    # Hvis vi har et revisionsbogstav i part number, brug det
                    base_part, rev_letter = self.extract_part_number(targetFileNames[i])
                    if rev_letter:
                        revList[i] = rev_letter
                    else:
                        revList[i] = revChar

            # Opdater REV kolonne i BOM fanen hvis den blev fundet
            if c_rev:
                for i, rev in enumerate(revList, start=2):
                    source_sheet.cell(row=i, column=c_rev).value = rev

            colQTY = df.columns.get_loc('QTY')
            df.insert(colQTY+1, column='Total QTY', value=totalQTY)

            # Initialiser CategoryParser
            parser = CategoryParser()
            logger.info("Starting categorization process")

            # Gruppér data efter kategori, men kun for rækker der skal inkluderes
            categorized_data = {}
            piping_parents = set()  # Gem alle piping parent items
            processed_basic_components = set()  # Hold styr på hvilke Basic Components vi har set
            
            # Først find alle piping parent items og Basic Components
            for idx, row in df.iterrows():
                if pd.notna(row['Part Number']) and self.should_include_row(row, df):
                    part_number = str(row['Part Number'])
                    logger.info(f"Processing part number: {part_number}")
                    category, type_ = parser.categorize(part_number)
                    logger.info(f"Category: {category}, Type: {type_}")
                    
                    # Tjek om det er en Basic Component eller Project Specific
                    if part_number.startswith('0000-3') or (part_number.startswith('0000-') and 'PS' in part_number):
                        # Tilføj til den korrekte kategori hvis vi ikke har set den før
                        target_category = 'Basic Components' if part_number.startswith('0000-3') else 'Project Specific'
                        if part_number not in processed_basic_components:
                            if target_category not in categorized_data:
                                categorized_data[target_category] = {'rows': [], 'indices': [], 'types': []}
                            categorized_data[target_category]['rows'].append(row)
                            categorized_data[target_category]['indices'].append(idx + 2)
                            categorized_data[target_category]['types'].append(type_)
                            processed_basic_components.add(part_number)
                            logger.info(f"Added to {target_category} category: {part_number}")
                    
                    if category == 'Piping':
                        item_number = str(row['Item'])
                        if '.' not in item_number:  # Kun top-level items
                            piping_parents.add(item_number)
                            logger.info(f"Added piping parent: {item_number}")
            
            # Derefter samle data med speciel håndtering af piping
            for idx, row in df.iterrows():
                if pd.notna(row['Part Number']) and self.should_include_row(row, df):
                    part_number = str(row['Part Number'])
                    logger.info(f"Processing part number for categorization: {part_number}")
                    category, type_ = parser.categorize(part_number)
                    logger.info(f"Category: {category}, Type: {type_}")
                    
                    item_number = str(row['Item'])
                    
                    # Tjek om dette er en child af en piping parent
                    is_piping_child = False
                    for parent in piping_parents:
                        if item_number.startswith(parent + '.'):
                            is_piping_child = True
                            logger.info(f"Item {item_number} is child of piping parent {parent}")
                            break
                    
                    # Hvis det er en piping eller en child af en piping, tilføj til piping kategorien
                    if category == 'Piping' or is_piping_child:
                        if 'Piping' not in categorized_data:
                            categorized_data['Piping'] = {'rows': [], 'indices': [], 'types': []}
                        categorized_data['Piping']['rows'].append(row)
                        categorized_data['Piping']['indices'].append(idx + 2)
                        categorized_data['Piping']['types'].append(type_)
                        logger.info(f"Added to Piping category: {part_number}")
                    # Tilføj kun til den oprindelige kategori hvis det ikke er en piping eller piping child
                    # OG det ikke er en Basic Component eller Project Specific vi allerede har set
                    elif not is_piping_child and part_number not in processed_basic_components:
                        if category not in categorized_data:
                            categorized_data[category] = {'rows': [], 'indices': [], 'types': []}
                        categorized_data[category]['rows'].append(row)
                        categorized_data[category]['indices'].append(idx + 2)
                        categorized_data[category]['types'].append(type_)
                        logger.info(f"Added to {category} category: {part_number}")

            # Opret nye sheets for hver kategori
            for category, data in categorized_data.items():
                if data['rows']:
                    # Opret ny sheet
                    ws = wb.create_sheet(title=category)
                    
                    # Kopier header fra source sheet
                    for col in range(1, source_sheet.max_column + 1):
                        ws.cell(row=1, column=col).value = source_sheet.cell(row=1, column=col).value
                    
                    # Kopier data og Total QTY
                    for i, row_idx in enumerate(data['indices']):
                        for col in range(1, source_sheet.max_column + 1):
                            ws.cell(row=i+2, column=col).value = source_sheet.cell(row=row_idx, column=col).value
                        
                        # Find eller opret Type kolonne
                        type_col = None
                        for col in range(1, source_sheet.max_column + 1):
                            if ws.cell(row=1, column=col).value == 'Type':
                                type_col = col
                                break
                        
                        # Hvis Type kolonne ikke findes, søg efter Keywords
                        if type_col is None:
                            for col in range(1, source_sheet.max_column + 1):
                                if ws.cell(row=1, column=col).value == 'Keywords':
                                    type_col = col
                                    ws.cell(row=1, column=col).value = 'Type'
                                    break
                        
                        # Hvis ingen Type eller Keywords kolonne findes, opret en ny
                        if type_col is None:
                            type_col = source_sheet.max_column + 1
                            ws.cell(row=1, column=type_col).value = 'Type'
                        
                        # Indsæt typen
                        ws.cell(row=i+2, column=type_col).value = data['types'][i]

                    # Opret Drawing kolonne lige efter Type kolonnen
                    drawing_col = type_col + 1  # Placer Drawing kolonnen lige efter Type kolonnen
                    ws.cell(row=1, column=drawing_col).value = 'Drawing'

                    # Juster kolonnebredder og formater header-rækken
                    self.adjust_column_widths(ws)
                    
                    # Hvis det er Piping fanen, tilføj gruppering
                    if category == 'Piping':
                        self.group_rows_by_level(ws)

                    # Opret kategorimappe
                    category_dir = os.path.join(dest_path, category)
                    os.makedirs(category_dir, exist_ok=True)

                    # Kopier relevante filer til kategorimappen og opret links
                    for i, row in enumerate(data['rows']):
                        part_number = str(row['Part Number'])
                        if pd.notna(part_number):
                            try:
                                file_idx = targetFileNames.index(part_number)
                                if file_idx < len(fileDestinations):
                                    # Få den korrekte kategori fra CategoryParser
                                    actual_category, _ = parser.categorize(part_number)
                                    actual_category_dir = os.path.join(dest_path, actual_category)
                                    os.makedirs(actual_category_dir, exist_ok=True)
                                    
                                    pdf_files = [f for f in fileDestinations[file_idx] if f.endswith('.pdf')]
                                    for pdf_file in pdf_files:
                                        dest_pdf = os.path.join(actual_category_dir, os.path.basename(pdf_file))
                                        shutil.copy2(pdf_file, dest_pdf)
                                        
                                        # Opret hyperlink i Excel med korrekt relativ sti
                                        relative_path = os.path.join(actual_category, os.path.basename(pdf_file))
                                        ws.cell(row=i+2, column=drawing_col).value = f'=HYPERLINK("{relative_path}","PDF")'
                                        ws.cell(row=i+2, column=drawing_col).style = 'Hyperlink'

                                    # Find og kopier DWG filer hvis inkluderet
                                    if self.include_dwg.get():
                                        dwg_files = [f for f in fileDestinations[file_idx] if f.endswith('.dwg')]
                                        for dwg_file in dwg_files:
                                            dest_dwg = os.path.join(actual_category_dir, os.path.basename(dwg_file))
                                            shutil.copy2(dwg_file, dest_dwg)
                            except ValueError:
                                logger.warning(f"Part number {part_number} not found in targetFileNames")
                                ws.cell(row=i+2, column=drawing_col).value = "Not available"

            # Kopier root files til rodmappen
            for i, part_number in enumerate(targetFileNames):
                if part_number in root_files:  # Tjek om det er en root file
                    if i < len(fileDestinations):
                        # Kopier PDF filer
                        pdf_files = [f for f in fileDestinations[i] if f.endswith('.pdf')]
                        for pdf_file in pdf_files:
                            dest_pdf = os.path.join(dest_path, os.path.basename(pdf_file))  # Kopier direkte til rodmappen
                            shutil.copy2(pdf_file, dest_pdf)
                            logger.info(f"Copied root PDF file to root directory: {dest_pdf}")
                        
                        # Kopier DWG filer hvis inkluderet
                        if self.include_dwg.get():
                            dwg_files = [f for f in fileDestinations[i] if f.endswith('.dwg')]
                            for dwg_file in dwg_files:
                                dest_dwg = os.path.join(dest_path, os.path.basename(dwg_file))  # Kopier direkte til rodmappen
                                shutil.copy2(dwg_file, dest_dwg)
                                logger.info(f"Copied root DWG file to root directory: {dest_dwg}")

            # Efter al kategorisering og filkopiering
            # Opret Partlist fanen
            if 'Partlist' in wb.sheetnames:
                wb.remove(wb['Partlist'])
            partlist = wb.create_sheet(title='Partlist')
            
            # Find alle kolonner fra BOM fanen, undtagen dem vi ikke vil have med
            exclude_columns = ['Item', 'QTY', 'Total QTY', 'Keywords']
            headers = []
            header_positions = {}
            
            for col in range(1, source_sheet.max_column + 1):
                header = source_sheet.cell(row=1, column=col).value
                if header and header not in exclude_columns:
                    headers.append(header)
                    header_positions[header] = len(headers)  # 1-baseret position i Partlist
                    partlist.cell(row=1, column=len(headers)).value = header
            
            # Tilføj Type kolonne hvis den ikke allerede findes
            if 'Type' not in headers:
                headers.append('Type')
                header_positions['Type'] = len(headers)
                partlist.cell(row=1, column=len(headers)).value = 'Type'
            
            # Tilføj Total QTY som sidste kolonne
            headers.append('Total QTY')
            header_positions['Total QTY'] = len(headers)
            partlist.cell(row=1, column=len(headers)).value = 'Total QTY'
            
            # Opbyg DataFrame med kun de relevante kolonner
            data = []
            for row in range(2, source_sheet.max_row + 1):
                row_data = {}
                # Tjek BOM Structure
                bom_structure = str(source_sheet.cell(row=row, column=df.columns.get_loc('BOM Structure') + 1).value).strip().upper()
                item_number = str(source_sheet.cell(row=row, column=df.columns.get_loc('Item') + 1).value)
                
                # Skip hvis rækken er Phantom
                if bom_structure == 'PHANTOM':
                    continue
                    
                # Tjek om rækken er child af en Inseparable
                is_child_of_inseparable = False
                if '.' in item_number:
                    parts = item_number.split('.')
                    for i in range(len(parts)-1, 0, -1):
                        parent_number = '.'.join(parts[:i])
                        for parent_row in range(2, source_sheet.max_row + 1):
                            if str(source_sheet.cell(row=parent_row, column=df.columns.get_loc('Item') + 1).value) == parent_number:
                                parent_structure = str(source_sheet.cell(row=parent_row, column=df.columns.get_loc('BOM Structure') + 1).value).strip().upper()
                                if parent_structure == 'INSEPARABLE':
                                    is_child_of_inseparable = True
                                    break
                        if is_child_of_inseparable:
                            break
                
                if is_child_of_inseparable:
                    continue
                
                # Tilføj data fra relevante kolonner
                for header in headers:  # Inkluder alle headers inklusive Type
                    col_idx = None
                    # Find kolonneindeks for header i source_sheet
                    for col in range(1, source_sheet.max_column + 1):
                        cell_value = source_sheet.cell(row=1, column=col).value
                        if cell_value == header:
                            col_idx = col
                            break
                    
                    if col_idx:
                        value = source_sheet.cell(row=row, column=col_idx).value
                        row_data[header] = value
                    elif header == 'Type':  # Hvis det er Type kolonnen
                        part_number = source_sheet.cell(row=row, column=df.columns.get_loc('Part Number') + 1).value
                        if part_number:
                            category, type_ = parser.categorize(str(part_number))
                            row_data['Type'] = type_
                
                # Tilføj Total QTY
                total_qty = source_sheet.cell(row=row, column=df.columns.get_loc('Total QTY') + 1).value
                try:
                    row_data['Total QTY'] = float(total_qty) if total_qty is not None else 0
                except (ValueError, TypeError):
                    row_data['Total QTY'] = 0
                
                data.append(row_data)
            
            # Konverter til DataFrame og gruppér
            df_partlist = pd.DataFrame(data)
            if not df_partlist.empty:
                # Gruppér efter Part Number, REV og Description, sumér Total QTY
                group_cols = ['Part Number', 'REV', 'Description']
                agg_dict = {'Total QTY': 'sum'}
                # Tilføj 'first' aggregering for alle andre kolonner
                for col in df_partlist.columns:
                    if col not in group_cols + ['Total QTY']:
                        agg_dict[col] = 'first'
                
                grouped = df_partlist.groupby(group_cols, as_index=False).agg(agg_dict)
                grouped = grouped.sort_values('Part Number')
                
                # Indsæt grupperet data i Partlist
                for idx, row in grouped.iterrows():
                    excel_row = idx + 2
                    for header in headers:
                        col_pos = header_positions[header]
                        value = row[header]
                        if pd.notna(value):
                            partlist.cell(row=excel_row, column=col_pos).value = value
            
            # Juster kolonnebredder og formater header-rækken for Partlist
            self.adjust_column_widths(partlist)

            # Gem workbook med nyt navn
            output_filename = f"{new_folder_name} - BOM.xlsx"
            output_path = os.path.join(dest_path, output_filename)
            wb.save(output_path)

            # Efter al filkopiering og processering
            self.update_progress(90, "Checking for missing DWG files...")
            
            # Tjek for manglende DWG filer med samme revision
            missing_dwg_info = self.check_matching_revisions()
            
            self.update_progress(95, "Cleaning up empty folders...")
            self.remove_empty_folders(dest_path)
            
            duration = (time.time() - start_time)
            
            # Vis den samlede slut-dialog
            self.show_final_dialog(duration, missing_dwg_info, dest_path)
            
            # Genaktiver start-knappen
            self.start_button.config(state='normal')
            
        except Exception as e:
            logger.error(f"Error in process_bom: {str(e)}")
            messagebox.showerror("Error", f"An error occurred: {str(e)}")
            # Genaktiver start-knappen ved fejl
            self.start_button.config(state='normal')

    def handle_bom_structure(self, sheet):
        """
        Håndterer BOM Structure regler:
        - Beholder rækker markeret som Inseparable og sletter kun deres direkte children
        - Beholder rækker med part numbers, der starter med 0000-3, og sletter kun deres direkte children
        - Sletter Phantom items
        """
        max_row = sheet.UsedRange.Rows.Count
        rows_to_delete = set()
        hierarchy = {}  # item_number -> [start_row, level]

        logger.info(f"Starting BOM structure handling. Total rows: {max_row}")

        # Opbyg hierarkisk struktur
        for row in range(2, max_row + 1):
            item_number = str(sheet.Cells(row, 1).Value).strip()  # Item i kolonne A
            if not item_number:
                continue

            # Bestem niveau baseret på antal punktummer
            level = len(item_number.split('.'))

            # Gem dette item
            hierarchy[item_number] = [row, level]

        # Find rækker der skal slettes
        for item_number, (start_row, level) in hierarchy.items():
            structure_type = str(sheet.Cells(start_row, 6).Value).strip()  # BOM Structure i kolonne F
            part_number = str(sheet.Cells(start_row, 2).Value).strip()     # Part Number i kolonne B

            # Håndter Phantom items
            if structure_type == "Phantom":
                rows_to_delete.add(start_row)
                logger.info(f"Marked Phantom row {start_row} for deletion")

            # Håndter Inseparable items
            elif structure_type == "Inseparable":
                # Behold Inseparable-rækken, men slet kun dens direkte children
                for child_item, (child_start, child_level) in hierarchy.items():
                    if child_item.startswith(item_number + '.') and child_level == level + 1:
                        rows_to_delete.add(child_start)
                        logger.info(f"Marked direct child row {child_start} of {item_number} for deletion")

            # Håndter 0000-3 part numbers
            elif part_number.startswith('0000-3'):
                # Behold 0000-3-rækken, men slet kun dens direkte children
                for child_item, (child_start, child_level) in hierarchy.items():
                    if child_item.startswith(item_number + '.') and child_level == level + 1:
                        rows_to_delete.add(child_start)
                        logger.info(f"Marked direct child row {child_start} of {item_number} for deletion")

        logger.info(f"Total rows marked for deletion: {len(rows_to_delete)}")
        if rows_to_delete:
            logger.info(f"Rows to delete: {sorted(rows_to_delete)}")

        # Slet rækker fra bunden og op
        for row in sorted(rows_to_delete, reverse=True):
            try:
                sheet.Rows(row).Delete()
                logger.info(f"Successfully deleted row {row}")
            except Exception as e:
                logger.error(f"Error deleting row {row}: {str(e)}")

    def check_database_status(self):
        """Tjekker hvornår databasen sidst blev opdateret"""
        try:
            db_path = r'\\192.168.170.18\drawings\file_index.db'
            if os.path.exists(db_path):
                conn = sqlite3.connect(db_path)
                cursor = conn.cursor()
                cursor.execute("SELECT value FROM metadata WHERE key = 'last_scan_time'")
                result = cursor.fetchone()
                conn.close()
                
                if result:
                    last_update = datetime.fromtimestamp(float(result[0]))
                    self.db_status.set(f"Database last updated: {last_update.strftime('%d/%m-%Y - %H:%M')}")
                else:
                    self.db_status.set("Database status: Never updated")
            else:
                self.db_status.set("Database not found")
        except Exception as e:
            logger.error(f"Error checking database status: {e}")
            self.db_status.set("Could not check database status")

    def get_latest_revision(self, files, part_number):
        """Find den seneste revision af en fil"""
        if not files:
            return None
            
        latest_file = None
        latest_rev = ''
        
        for file in files:
            filename = os.path.basename(file)
            # Find position efter part number
            pos = filename.find(part_number) + len(part_number)
            if pos < len(filename):
                # Tag det næste tegn som revision (efter bindestreg)
                if pos + 1 < len(filename) and filename[pos] == '-':
                    rev = filename[pos + 1]
                    if rev > latest_rev or not latest_rev:
                        latest_rev = rev
                        latest_file = file
                        
        # Hvis ingen revision blev fundet, brug den første fil
        if not latest_file and files:
            latest_file = files[0]
            
        return latest_file

    def get_file_info(self, filepath, part_number):
        """Henter revision og oprettelsesdato for en fil"""
        filename = os.path.basename(filepath)
        logger.info(f"Getting file info for: {filename}")
        
        # Find position efter part number
        pos = filename.find(part_number) + len(part_number)
        revision = ''
        if pos < len(filename) and filename[pos] == '-' and pos + 1 < len(filename):
            revision = filename[pos + 1]
            logger.info(f"Found revision: {revision}")
            
        # Få filens oprettelsesdato (uden klokkeslæt)
        creation_time = datetime.fromtimestamp(os.path.getctime(filepath))
        creation_date = creation_time.date()
        logger.info(f"File creation date: {creation_date}")
        
        return revision, creation_date

    def find_matching_files(self, files_by_type, part_number):
        """Finder PDF og DWG filer der matcher i revision og dato"""
        logger.info(f"Searching for matching files for part number: {part_number}")
        
        # Filtrer FOR REVIEW filer fra
        pdf_files = [f for f in files_by_type.get('pdf', []) if not f.upper().endswith('_FOR REVIEW.PDF')]
        if not pdf_files:
            logger.warning(f"No PDF files found for {part_number}")
            return None, None
            
        # Find den seneste PDF revision
        latest_pdf = None
        latest_rev = ''
        for pdf_file in pdf_files:
            # Tjek om filen starter med part number
            pdf_name = os.path.basename(pdf_file).lower()
            if pdf_name.startswith(part_number.lower()):
                latest_pdf = pdf_file
                logger.info(f"Found matching PDF file: {pdf_file}")
                break
        
        if not latest_pdf:
            logger.warning(f"No valid PDF found for {part_number}")
            return None, None
            
        # Tjek om der findes en tilsvarende DWG fil
        dwg_file = latest_pdf[:-4] + '.dwg'  # Erstat .pdf med .dwg
        logger.info(f"Checking for DWG file at: {dwg_file}")
        
        if os.path.exists(dwg_file):
            logger.info(f"Found matching DWG file: {dwg_file}")
            return latest_pdf, dwg_file
        else:
            logger.warning(f"No matching DWG file found at: {dwg_file}")
            return latest_pdf, None

    def extract_part_number(self, part_number):
        """Extracts the base part number and revision letter if present"""
        # Check if the last part is a single letter after a hyphen
        parts = part_number.split('-')
        if len(parts) > 1 and len(parts[-1]) == 1 and parts[-1].isalpha():
            # Return both the base part number and the revision letter
            return '-'.join(parts[:-1]), parts[-1]
        return part_number, ''

    def show_final_dialog(self, duration, missing_dwg_info, dest_path):
        """Viser en dialog med resultater og mulighed for at gemme part numbers"""
        dialog = tk.Toplevel(self.root)
        dialog.title("Processing Complete")
        
        # Sæt vinduesstørrelse
        window_width = 600
        window_height = 400
        
        # Beregn position for at centrere vinduet
        screen_width = dialog.winfo_screenwidth()
        screen_height = dialog.winfo_screenheight()
        x = (screen_width - window_width) // 2
        y = (screen_height - window_height) // 2
        
        # Sæt vinduesstørrelse og position
        dialog.geometry(f'{window_width}x{window_height}+{x}+{y}')
        
        dialog.transient(self.root)
        dialog.grab_set()
        
        # Sæt dialog til at være øverst
        dialog.attributes('-topmost', True)
        
        # Hovedframe
        main_frame = ttk.Frame(dialog, padding="10")
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # Tid information
        time_label = ttk.Label(main_frame, text=f"Processing completed in {duration:.2f} seconds")
        time_label.pack(pady=5)
        
        # Treeview for manglende filer
        tree_frame = ttk.Frame(main_frame)
        tree_frame.pack(fill=tk.BOTH, expand=True, pady=10)
        
        # Opret Treeview
        tree = ttk.Treeview(tree_frame, columns=('part_number', 'status'), show='headings')
        tree.heading('part_number', text='Part Number')
        tree.heading('status', text='Status')
        
        # Sæt kolonnebredder
        tree.column('part_number', width=200)
        tree.column('status', width=300)
        
        # Tilføj scrollbar
        scrollbar = ttk.Scrollbar(tree_frame, orient=tk.VERTICAL, command=tree.yview)
        tree.configure(yscrollcommand=scrollbar.set)
        
        tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # Dictionary til at holde styr på status for hvert part number
        part_status = {}

        # Byg lookup for Item -> Part Number
        item_to_partnumber = {}
        # Byg lookup for Part Number -> Item (kan være flere items med samme part number)
        partnumber_to_items = {}
        excel_files = [f for f in os.listdir(dest_path) if f.lower().endswith('.xlsx')]
        if not excel_files:
            messagebox.showerror("Error", "No Excel file found in destination folder.")
            return
        wb = openpyxl.load_workbook(os.path.join(dest_path, excel_files[0]))
        if 'BOM' in wb.sheetnames:
            sheet = wb['BOM']
            item_col = None
            bomstruct_col = None
            partnum_col = None
            for col in range(1, sheet.max_column + 1):
                if sheet.cell(row=1, column=col).value == 'Item':
                    item_col = col
                if sheet.cell(row=1, column=col).value == 'BOM Structure':
                    bomstruct_col = col
                if sheet.cell(row=1, column=col).value == 'Part Number':
                    partnum_col = col
            if item_col and bomstruct_col and partnum_col:
                item_bomstructure = {}
                item_parent = {}
                for row in range(2, sheet.max_row + 1):
                    item = str(sheet.cell(row=row, column=item_col).value)
                    partnum = str(sheet.cell(row=row, column=partnum_col).value)
                    bomstruct = str(sheet.cell(row=row, column=bomstruct_col).value).strip().upper()
                    item_bomstructure[item] = bomstruct
                    item_to_partnumber[item] = partnum
                    if partnum not in partnumber_to_items:
                        partnumber_to_items[partnum] = []
                    partnumber_to_items[partnum].append(item)
                    if '.' in item:
                        parent = '.'.join(item.split('.')[:-1])
                        item_parent[item] = parent
                # Find alle items der er child til en inseparable parent
                inseparable_items = set()
                for item, struct in item_bomstructure.items():
                    if struct == 'INSEPARABLE':
                        # Markér alle children (direkte og indirekte)
                        stack = [item]
                        while stack:
                            parent = stack.pop()
                            for child, p in item_parent.items():
                                if p == parent and child not in inseparable_items:
                                    inseparable_items.add(child)
                                    stack.append(child)
                # Find alle part numbers der er child til inseparable parent
                inseparable_partnumbers = set()
                for item in inseparable_items:
                    pn = item_to_partnumber.get(item)
                    if pn:
                        inseparable_partnumbers.add(pn)

        for part_number in targetFileNames:
            # Hvis part_number er child af inseparable parent, så spring helt over (skal ikke vises i listen)
            if part_number in inseparable_partnumbers:
                continue
            pdf_found = False
            dwg_found = False

            # Tjek om det er en root file
            if part_number in root_files:
                # Kig kun i rodmappen
                for file in os.listdir(dest_path):
                    if file.startswith(part_number):
                        if file.endswith('.pdf'):
                            pdf_found = True
                        elif file.endswith('.dwg'):
                            dwg_found = True
            else:
                # Gennemgå alle mapper i destinationsmappen
                for category in os.listdir(dest_path):
                    category_path = os.path.join(dest_path, category)
                    if os.path.isdir(category_path):
                        # Tjek alle filer i kategorimappen
                        for file in os.listdir(category_path):
                            if file.startswith(part_number):
                                if file.endswith('.pdf'):
                                    pdf_found = True
                                elif file.endswith('.dwg'):
                                    dwg_found = True

            # Bestem status baseret på fundne filer
            if not pdf_found:
                part_status[part_number] = "PDF: No"
            elif not dwg_found and self.include_dwg.get():
                part_status[part_number] = "DWG: No"
        
        # Sorter problemer: Først efter om de starter med 0000-7, derefter efter part number
        def sort_key(item):
            part_number, status = item
            # Sæt 0000-7 til at komme til sidst
            if part_number.startswith('0000-7'):
                return (1, part_number)  # 1 for at sende til bunden
            return (0, part_number)  # 0 for at beholde øverst
        
        # Konverter dictionary til liste og sorter
        sorted_problems = sorted(part_status.items(), key=sort_key)
        
        # Tilføj sorterede problemer til treeview
        for part_number, status in sorted_problems:
            tree.insert('', 'end', values=(part_number, status))
        
        # Total tæller
        total_label = ttk.Label(main_frame, text=f"Total issues found: {len(sorted_problems)}")
        total_label.pack(pady=5)
        
        # Knapper
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(pady=10)
        
        if sorted_problems:
            save_button = ttk.Button(button_frame, text="Save Part Numbers", 
                                   command=lambda: self.save_part_numbers(sorted_problems))
            save_button.pack(side=tk.LEFT, padx=5)
        
        def open_folder():
            try:
                os.startfile(dest_path)
            except Exception as e:
                logger.error(f"Error opening folder: {e}")
                messagebox.showerror("Error", f"Could not open folder: {e}")
        
        open_button = ttk.Button(button_frame, text="Open Destination Folder", 
                               command=open_folder)
        open_button.pack(side=tk.LEFT, padx=5)
        
        # Ændret til kun at lukke dialogen
        close_button = ttk.Button(button_frame, text="Close", 
                               command=dialog.destroy)
        close_button.pack(side=tk.LEFT, padx=5)
        
        # Håndter lukning af vinduet (X knappen)
        dialog.protocol("WM_DELETE_WINDOW", dialog.destroy)

    def save_part_numbers(self, problems):
        """Gemmer part numbers med problemer til en tekstfil"""
        desktop = os.path.join(os.path.expanduser("~"), "Desktop")
        filename = os.path.join(desktop, f"missing_files_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt")
        
        with open(filename, 'w', encoding='utf-8') as f:
            f.write(f"Part Numbers with missing files\n")
            f.write(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n")
            
            for part_number, status in problems:
                f.write(f"Part Number: {part_number}\n")
                f.write(f"Status: {status}\n\n")
        
        messagebox.showinfo("Saved", f"File saved as:\n{filename}")

    def check_matching_revisions(self):
        """Tjekker om der er matchende revisioner mellem PDF og DWG filer"""
        missing_dwg_info = {}
        
        for idx, part_number in enumerate(targetFileNames):
            if idx < len(fileDestinations):
                pdf_files = [f for f in fileDestinations[idx] 
                           if f.lower().endswith('.pdf') and 
                           not f.upper().endswith('_FOR REVIEW.PDF')]
                dwg_files = [f for f in fileDestinations[idx] 
                           if f.lower().endswith('.dwg')]
                
                if pdf_files:  # Hvis der er en PDF fil
                    pdf_file = pdf_files[0]  # Vi har allerede den seneste revision
                    pdf_rev, pdf_date = self.get_file_info(pdf_file, part_number)
                    
                    matching_dwg = None
                    if dwg_files:
                        for dwg_file in dwg_files:
                            dwg_rev, dwg_date = self.get_file_info(dwg_file, part_number)
                            if dwg_rev == pdf_rev and dwg_date == pdf_date:
                                matching_dwg = dwg_file
                                break
                    
                    if not matching_dwg:  # Hvis ingen matchende DWG blev fundet
                        missing_dwg_info[part_number] = {
                            'pdf': pdf_file,
                            'dwg': dwg_files[0] if dwg_files else None
                        }
        
        return missing_dwg_info

    def get_file_revision(self, filepath, part_number):
        """Henter revisionsbogstavet fra et filnavn"""
        filename = os.path.basename(filepath)
        pos = filename.find(part_number) + len(part_number)
        if pos < len(filename) and filename[pos] == '-' and pos + 1 < len(filename):
            return filename[pos + 1]
        return ''

    def search_files_in_database(self, part_numbers):
        """Søger efter filer i databasen for alle part numbers"""
        try:
            db_path = r'\\192.168.170.18\drawings\file_index.db'
            if not os.path.exists(db_path):
                logger.error("Database file not found")
                return False

            conn = sqlite3.connect(db_path)
            cursor = conn.cursor()
            
            # Nulstil fileDestinations
            global fileDestinations
            fileDestinations = [[] for _ in range(len(part_numbers))]
            
            # Søg efter hver part number
            for i, part_number in enumerate(part_numbers):
                files_by_type = {'pdf': [], 'dwg': []}
                
                logger.info(f"Searching for files with part number: {part_number}")
                
                for ext in ['.pdf', '.dwg']:
                    # Søg efter filer der starter med part number
                    search_pattern = f"{part_number}%"
                    
                    cursor.execute("""
                        SELECT path FROM files 
                        WHERE filename LIKE ? AND file_type = ?
                        ORDER BY modified_time DESC
                    """, (search_pattern, ext))
                    
                    files = [row[0] for row in cursor.fetchall()]
                    logger.info(f"Found {len(files)} {ext} files for pattern {search_pattern}")
                    files_by_type[ext[1:]].extend(files)
                
                # Find matchende filer
                pdf_file, dwg_file = self.find_matching_files(files_by_type, part_number)
                if pdf_file:
                    fileDestinations[i].append(pdf_file)
                    if dwg_file:
                        fileDestinations[i].append(dwg_file)
                        logger.info(f"Found matching PDF and DWG for {part_number}")
                    else:
                        logger.info(f"Found only PDF for {part_number}")
            
            conn.close()
            return True
            
        except Exception as e:
            logger.error(f"Error searching database: {e}")
            return False

    def search_files_directly(self, source_dir):
        """Søger efter filer direkte i mapperne"""
        try:
            # Nulstil fileDestinations
            global fileDestinations, targetFileNames
            fileDestinations = [[] for _ in range(len(targetFileNames))]
            
            # Dictionary til at samle filer efter type
            files_by_part = {}
            
            logger.info(f"Starting direct search in directory: {source_dir}")
            
            # Start concurrent søgning for at samle alle filer
            with ThreadPoolExecutor() as executor:
                futures = [executor.submit(self.scan_directory_task, entry, files_by_part) 
                          for entry in os.scandir(source_dir)]
                for future in futures:
                    future.result()
            
            # Find matchende filer for hvert part number
            for idx, part_number in enumerate(targetFileNames):
                logger.info(f"Processing part number: {part_number}")
                
                # Find alle filer der matcher part number
                all_pdf_files = []
                all_dwg_files = []
                
                if part_number in files_by_part:
                    all_pdf_files.extend(files_by_part[part_number]['pdf'])
                    all_dwg_files.extend(files_by_part[part_number]['dwg'])
                    logger.info(f"Found {len(all_pdf_files)} PDF files and {len(all_dwg_files)} DWG files for {part_number}")
                
                # Find matchende filer
                pdf_file, dwg_file = self.find_matching_files(
                    {'pdf': all_pdf_files, 'dwg': all_dwg_files}, 
                    part_number
                )
                
                if pdf_file:
                    fileDestinations[idx].append(pdf_file)
                    if dwg_file:
                        fileDestinations[idx].append(dwg_file)
                        logger.info(f"Found matching PDF and DWG for {part_number}")
                    else:
                        logger.info(f"Found only PDF for {part_number}")
            
            return True
            
        except Exception as e:
            logger.error(f"Error in direct file search: {e}")
            return False

    def scan_directory_task(self, entry, files_by_part):
        """Scanner en enkelt mappe eller fil og gemmer alle fundne filer i files_by_part"""
        try:
            if entry.is_dir(follow_symlinks=False):
                file_paths = []
                for sub_entry in os.scandir(entry.path):
                    file_paths.extend(self.scan_directory_task(sub_entry, files_by_part))
                return file_paths
            elif entry.is_file(follow_symlinks=False):
                name = entry.name.lower()
                if '_for review.pdf' in name:  # Skip FOR REVIEW filer
                    return []
                    
                for targetName in targetFileNames:
                    # Tjek om filen starter med part number
                    if name.startswith(targetName.lower()):
                        # Gem filen hvis den er nyere end den nuværende
                        if name.endswith('.pdf'):
                            if targetName not in files_by_part:
                                files_by_part[targetName] = {'pdf': [], 'dwg': []}
                            if not files_by_part[targetName]['pdf']:
                                files_by_part[targetName]['pdf'].append(entry.path)
                                logger.info(f"Found PDF file: {entry.path} for part number {targetName}")
                            else:
                                # Tjek om den nye fil er nyere
                                current_time = os.path.getmtime(entry.path)
                                existing_time = os.path.getmtime(files_by_part[targetName]['pdf'][0])
                                if current_time > existing_time:
                                    files_by_part[targetName]['pdf'][0] = entry.path
                                    logger.info(f"Found newer PDF file: {entry.path} for part number {targetName}")
                        elif name.endswith('.dwg'):
                            if targetName not in files_by_part:
                                files_by_part[targetName] = {'pdf': [], 'dwg': []}
                            if not files_by_part[targetName]['dwg']:
                                files_by_part[targetName]['dwg'].append(entry.path)
                                logger.info(f"Found DWG file: {entry.path} for part number {targetName}")
                            else:
                                # Tjek om den nye fil er nyere
                                current_time = os.path.getmtime(entry.path)
                                existing_time = os.path.getmtime(files_by_part[targetName]['dwg'][0])
                                if current_time > existing_time:
                                    files_by_part[targetName]['dwg'][0] = entry.path
                                    logger.info(f"Found newer DWG file: {entry.path} for part number {targetName}")
                return [entry.path]
            return []
        except Exception as e:
            logger.error(f"Error in scan_directory_task: {e}")
            return []

def main():
    gui = ExcelCopyBOMGUI()
    gui.root.mainloop()

if __name__ == "__main__":
    main()

    #pyinstaller --onefile --noconsole --hidden-import=win32com.client --hidden-import=pythoncom --hidden-import=pandas --hidden-import=openpyxl --hidden-import=sqlite3 "Python/Cursor_ExcelCopyBOM/CursorCopyBOM.py"