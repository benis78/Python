import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
import tkinter as tk
from tkinter import filedialog, ttk  # Add ttk import
import os
import logging

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    filename='comparison.log'
)

class FileSelectionGUI:
    def __init__(self):
        self.old_file = ""
        self.new_file = ""
        
        self.root = tk.Tk()
        self.root.title("Excel Comparison Tool")
        self.root.geometry("600x250")  # Increased height for checkbox
        
        # Pass root as master to BooleanVar
        self.exclude_supplier_parts = tk.BooleanVar(master=self.root, value=True)
        
        # Old file selection
        tk.Label(self.root, text="Old Revision:").grid(row=0, column=0, padx=5, pady=5)
        self.old_path = tk.Entry(self.root, width=50)
        self.old_path.grid(row=0, column=1, padx=5, pady=5)
        tk.Button(self.root, text="Browse", command=self.browse_old).grid(row=0, column=2, padx=5, pady=5)
        
        # New file selection
        tk.Label(self.root, text="New Revision:").grid(row=1, column=0, padx=5, pady=5)
        self.new_path = tk.Entry(self.root, width=50)
        self.new_path.grid(row=1, column=1, padx=5, pady=5)
        tk.Button(self.root, text="Browse", command=self.browse_new).grid(row=1, column=2, padx=5, pady=5)
        
        # Add checkbox for supplier parts
        ttk.Checkbutton(
            self.root, 
            text="Exclude supplier parts (0000-700, 0000-701, 0000-702)",
            variable=self.exclude_supplier_parts
        ).grid(row=2, column=0, columnspan=3, pady=10)
        
        # Move Go button to row 3
        tk.Button(self.root, text="Go", command=self.start_comparison).grid(row=3, column=1, pady=20)
        
        self.root.mainloop()
    
    def browse_old(self):
        filename = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        self.old_path.delete(0, tk.END)
        self.old_path.insert(0, filename)
        self.old_file = filename
    
    def browse_new(self):
        filename = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        self.new_path.delete(0, tk.END)
        self.new_path.insert(0, filename)
        self.new_file = filename
    
    def start_comparison(self):
        if self.old_file and self.new_file:
            self.root.quit()

# Add filter function
def filter_supplier_parts(df):
    """Filter out supplier parts from DataFrame"""
    supplier_patterns = ['0000-700', '0000-701', '0000-702']
    mask = ~df['Part Number'].str.startswith(tuple(supplier_patterns), na=False)
    return df[mask]

def main():
    # Start GUI to get file paths
    gui = FileSelectionGUI()
    old_file_path = gui.old_file
    new_file_path = gui.new_file
    exclude_supplier = gui.exclude_supplier_parts.get()  # Get checkbox value
    
    if not old_file_path or not new_file_path:
        return
    
    # Get output directory (same as new file)
    output_dir = os.path.dirname(new_file_path)
    result_path = os.path.join(output_dir, "BOM_Comparison_Composite_Result.xlsx")
    
    # Load and filter Excel files
    old_df = pd.read_excel(old_file_path)
    new_df = pd.read_excel(new_file_path)
    
    if exclude_supplier:
        old_df = filter_supplier_parts(old_df)
        new_df = filter_supplier_parts(new_df)
        logging.info("Supplier parts filtered out")

    # Definer alle kolonner (antager at begge filer har samme kolonner)
    columns_all = new_df.columns.tolist()

    # Opret dictionaries med composite nøgle: (Part Number, Description)
    def create_keyed_dict(df):
        keyed = {}
        for _, row in df.iterrows():
            key = (row["Part Number"], row["Description"])
            # Hvis der er flere rækker med samme composite nøgle, gemmes de i en liste
            if key in keyed:
                keyed[key].append(row)
            else:
                keyed[key] = [row]
        return keyed

    old_keyed = create_keyed_dict(old_df)
    new_keyed = create_keyed_dict(new_df)

    # Saml alle unikke composite nøgler
    all_keys = set(list(old_keyed.keys()) + list(new_keyed.keys()))

    # Definer formateringer
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    red_fill = PatternFill(start_color="FFCECE", end_color="FFCECE", fill_type="solid")  # New red background
    green_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")

    # Opret et nyt workbook til resultatet
    wb = Workbook()
    ws = wb.active
    ws.title = "Comparison Result"

    # Tilføj overskrift (kolonne navne) og status kolonne
    columns_with_status = ["Status"] + columns_all
    ws.append(columns_with_status)

    # Define columns to check for changes
    columns_to_check = ["REV", "Material", "QTY", "Total QTY", "D", "t", "L"]
    
    # Gå igennem alle nøgler
    for key in all_keys:
        old_rows = old_keyed.get(key, [])
        new_rows = new_keyed.get(key, [])
        has_changes = False

        if old_rows and new_rows:
            # For simplificering, sammenlignes kun den første række for nøglen i hver fil.
            old_row = old_rows[0]
            new_row = new_rows[0]
            result_row = []
            
            # Tjek om der er ændringer i kolonnerne
            for col in columns_all:
                old_val = old_row.get(col, "")
                new_val = new_row.get(col, "")
                if pd.isna(old_val): 
                    old_val = ""
                if pd.isna(new_val): 
                    new_val = ""
                
                if col in columns_to_check:
                    if old_val != new_val:
                        has_changes = True
                        cell_value = f"{new_val} (was: {old_val})"
                        result_row.append(cell_value)
                    else:
                        result_row.append(new_val)
                else:
                    # For columns not in columns_to_check, always use the new value without marking changes
                    result_row.append(new_val)
            
            if has_changes:
                ws.append(["Changed"] + result_row)
                current_row = ws.max_row
                for col_idx, col in enumerate(columns_all, start=2):
                    if "(was:" in str(ws.cell(row=current_row, column=col_idx).value):
                        ws.cell(row=current_row, column=col_idx).fill = yellow_fill

        elif new_rows and not old_rows:
            # Række findes kun i den nye fil: markér hele rækken med grøn baggrund.
            new_row = new_rows[0]
            result_row = [new_row.get(col, "") if not pd.isna(new_row.get(col, "")) else "" for col in columns_all]
            ws.append(["Added"] + result_row)
            current_row = ws.max_row
            for col_idx in range(2, len(columns_all) + 2):
                ws.cell(row=current_row, column=col_idx).fill = green_fill

        elif old_rows and not new_rows:
            # Række findes kun i den gamle fil: markér hele rækken med rød baggrund
            old_row = old_rows[0]
            result_row = [old_row.get(col, "") if not pd.isna(old_row.get(col, "")) else "" for col in columns_all]
            ws.append(["Removed"] + result_row)
            current_row = ws.max_row
            for col_idx in range(2, len(columns_all) + 2):
                ws.cell(row=current_row, column=col_idx).fill = red_fill  # Changed to red background

    # Gem resultatet
    wb.save(result_path)
    return result_path

if __name__ == "__main__":
    main()