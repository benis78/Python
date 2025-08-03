import openpyxl

# Åbn Excel-filen
wb = openpyxl.load_workbook('C:/Coding/Python/Cursor_ExcelCopyBOM/4003-02.1-A01-- - BOM.xlsx')
sheet = wb['BOM']

# Find Drawing-kolonnen
drawing_col = None
for col in range(1, sheet.max_column + 1):
    if sheet.cell(1, col).value == 'Drawing':
        drawing_col = col
        break

if drawing_col:
    print("Drawing kolonne fundet i kolonne", drawing_col)
    print("\nHyperlinks i Drawing kolonnen:")
    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, drawing_col)
        if cell.value:
            print(f"Række {row}: {cell.value}") 