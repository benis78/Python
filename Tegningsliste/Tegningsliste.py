import os
import tkinter as tk
import win32com
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from win32com import client
from tkinter import filedialog, messagebox
from time import gmtime, strftime


#start_time = time.time()

# os.chdir(ePath)
wb=Workbook()
#wb=openpyxl.load_workbook(filename='C:\\Users\\Jesper\\Desktop\\export_dataframe.xlsx')
#sheet=wb.sheetnames
sheet=wb.active

root = tk.Tk()
root.withdraw()

#t='Choose a Drawing folder'
#dest_path=filedialog.askdirectory(title=t.upper(), initialdir='\\\\192.168.170.9\\Bigadan\Bigadan\\1_PROJEKTER')
dest_path='C:\\Dropbox\\VizCon\\03-Teknisk dokumentation\\Vitek (1)\\Projekt'
destPath=(os.path.abspath(dest_path))
#destPath = (r'\\192.168.170.9\Bigadan\Bigadan\1_PROJEKTER\2500 Aabenraa\6.0 Drawings')

destDir=(os.path.basename)
# if dest_path == '':
#     os._exit(1)    
# print(os.path.split(os.path.abspath(os.path.join(destPath, os.pardir)))[1])
#print(destPath)
# Returnere parent folder der hvor projektnavnet er.

project=os.path.split(os.path.abspath(os.path.join(destPath, os.pardir)))[1]
projectnumber = project.split()[0]
# targetbase = projectnumber+'-4', projectnumber+'-6'
print(project)

# fileDestinations = [] # liste med samme længde som variablen targetFileNames, hvor hvert element til at starte med selv er en tom liste.
# filnavn=[] # liste med filnavnet
# created=[] # liste med dato for oprettelse

pdfDestinations=[]
pdfFilnavn=[]
pdfCreated = []
dwgDestinations=[]
targetExtensions = ['.pdf','.eprt']
targetDirs=['']

pdfExtension = '.pdf'
dwgExtension = '.dwg'
eprtExtension = '.eprt'
excludeExtension= 'sldprt','sldasm','slddrw'
#excludefolder='Old','old','OLD'

for root, dirs, files in os.walk(destPath,topdown=True):           # Vi looper igennem vores directory
    # if excludefolder in dirs:
    #     dirs.remove(excludefolder)
    if 'Old' in dirs:
        dirs.remove('Old')
    if 'old' in dirs:
        dirs.remove('old')
    if 'OLD' in dirs:
        dirs.remove('OLD')
    # print(dirs)
    for name in files:
        for targetExtension in targetExtensions:
            #print(targetExtension)
            if name.endswith(pdfExtension): 
                pdfDestinations.append(os.path.join(root,name))
                pdfFilnavn.append(os.path.splitext(name)[0])
                pdfCreated.append(strftime("%d/%m/%Y", gmtime(os.path.getctime(os.path.join(root,name)))))


# set col values
colDescription=1
colDato=2
colLink=3
# colEprtLink=4

# set row values
rowTitle=1
rowstart=2
bd = Side(style='thin', color="000000")

# Lock the all above and to the left of cell value
sheet.freeze_panes = sheet['A3']

sheet.auto_filter.ref = 'A2:C1500'

# Title
sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
sheet.cell(row=rowTitle, column=1).value = 'Tegningsliste'
sheet.cell(row=rowTitle, column=1).font = Font(size='16', bold=True)
sheet.cell(row=rowTitle, column=1).alignment = Alignment(horizontal='center')


# Link
sheet.cell(row=rowstart, column=colLink).value = 'Link'
sheet.cell(row=rowstart, column=colLink).font = Font(bold=True)
sheet.cell(row=rowstart, column=colLink).border = Border(bottom=bd, top=bd)
sheet.column_dimensions[get_column_letter(colLink)].width = '6.8'

# Date
sheet.cell(row=rowstart, column=colDato).value = 'Date'
sheet.cell(row=rowstart, column=colDato).font = Font(bold=True)
sheet.cell(row=rowstart, column=colDato).border = Border(bottom=bd, top=bd)
sheet.column_dimensions[get_column_letter(colDato)].width = '10.7'

#Description
sheet.cell(row=rowstart, column=colDescription).value ='Description'
sheet.cell(row=rowstart, column=colDescription).font = Font(bold=True)
sheet.cell(row=rowstart, column=colDescription).border = Border(bottom=bd, top=bd)
sheet.column_dimensions[get_column_letter(colDescription)].width = '71.5'

# # dwgLink
# sheet.cell(row=rowstart, column=colEprtLink).value = 'dwg'
# sheet.cell(row=rowstart, column=colEprtLink).font = Font(bold=True)
# sheet.cell(row=rowstart, column=colEprtLink).border = Border(bottom=bd, top=bd)
# sheet.column_dimensions[get_column_letter(colEprtLink)].width = '6.8'

i=rowstart+1
for File in pdfDestinations:
    sheet.cell(row=i, column=colLink).value = ('=HYPERLINK("'+File+'", "PDF")')
    sheet.cell(row=i, column=colLink).style = 'Hyperlink'               
    i+=1

i=rowstart+1
for des in pdfFilnavn:
    sheet.cell(row=i, column=colDescription).value = (des)
    i+=1

i=rowstart+1
for dato in pdfCreated:
    sheet.cell(row=i, column=colDato).value = (dato)
    sheet.cell(row=i, column=colDato).number_format = 'dd-mm-yyyy'
    i+=1

# i=rowstart+1
# for File in dwgDestinations:
#     sheet.cell(row=i, column=colEprtLink).value = ('=HYPERLINK("'+File+'", "dwg")')
#     sheet.cell(row=i, column=colEprtLink).style = 'Hyperlink'               
#     i+=1
#timestr = time.strftime("%Y%m%d-%H%M")


##Autofit column
# column_widths = []
# for row in sheet.iter_rows():
#     for i, cell in enumerate(row):
#         try:
#             column_widths[i] = max(column_widths[i], len(str(cell.value)))
#         except IndexError:
#             column_widths.append(len(str(cell.value)))

# for i, column_width in enumerate(column_widths):
#     sheet.column_dimensions[get_column_letter(i + 1)].width = column_width



#save excel   
sheet.title='Tegningsliste'    
#saveName=destPath+'\\'+project+' - List of drawings.xlsx'
saveName='C:\\Users\\Jesper Lund Hansen\\Desktop\\Tegningsliste\\Tegningsliste.xlsx'
wb.save(saveName)

yesno=tk.messagebox.askyesno(title='Tegningsliste', message='Vil du åbne filen i Excel?')

if yesno == 1:
    xl=win32com.client.Dispatch("Excel.Application")
    xl.Visible = True
    xl.Workbooks.Open(saveName)
else:
    tk.messagebox.showinfo(title='Tegningsliste', message='Færdig')


#Pyinstaller --onefile Tegningsliste.py