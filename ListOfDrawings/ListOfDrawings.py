import os
import tkinter as tk
import win32com.client as win32 
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from win32com import client
from tkinter import filedialog, messagebox
from time import gmtime, strftime
# from collections import OrderedDict
# import re



#start_time = time.time()

# os.chdir(ePath)
wb=Workbook()
#wb=openpyxl.load_workbook(filename='C:\\Users\\Jesper\\Desktop\\export_dataframe.xlsx')
#sheet=wb.sheetnames
sheet=wb.active

root = tk.Tk()
root.withdraw()

t='Choose a Drawing folder'
dest_path=filedialog.askdirectory(title=t.upper(), initialdir='\\\\192.168.170.9\\Bigadan\Bigadan\\1_PROJEKTER')
# dest_path='D:/Dropbox/Coding/BOM_New.xlsx'
destPath=(os.path.abspath(dest_path))
#destPath = (r'\\192.168.170.9\Bigadan\Bigadan\1_PROJEKTER\2500 Aabenraa\6.0 Drawings')

destDir=(os.path.basename)
if dest_path == '':
    os._exit(1)    
# print(os.path.split(os.path.abspath(os.path.join(destPath, os.pardir)))[1])
# print(destPath)
# Returnere parent folder der hvor projektnavnet er.

#project=os.path.split(os.path.abspath(os.path.join(destPath, os.pardir)))[1]
project=os.path.split(destPath.split()[0])[1]
#project="2500"
print(project)
#projectnumber = project.split()[0]
#print(projectnumber)
targetBase = project
# targetBase = projectnumber+'-0', projectnumber+'-4', projectnumber+'-6', projectnumber+'-8' # , '0000-3'


fileDestinations = [] # liste med samme længde som variablen targetFileNames, hvor hvert element til at starte med selv er en tom liste.
filnavn=[] # liste med filnavnet
created=[] # liste med dato for oprettelse

targetExtensions = ['.pdf']

excludeExtensions= ''#'_FOR REVIEW.pdf','_WIP.pdf','_Work In Progress.pdf'
#excludefolder='Old','old','OLD'

for root, dirs, files in os.walk(destPath,topdown=True):           # Vi looper igennem vores directory
    # if excludefolder in dirs:
    #     dirs.remove(excludefolder)
    if 'Old' in dirs:
        dirs.remove('Old')
    if 'old' in dirs:
        dirs.remove('old')
    if 'OLD' in dirs:
        dirs.remove('OLD')
    if 'Drawing Set' in dirs:
        dirs.remove('Drawing Set')
    if 'Drawing Sets' in dirs:
        dirs.remove('Drawing Sets')

    # print(dirs)
    for name in files:                              # Vi looper igennem hver fil i den aktuelle mappe
        for targetExtension in targetExtensions:
            if name.startswith(targetBase) and name.endswith(targetExtension):
            # if re.match(targetBase,name) and name.endswith(targetExtension): 
                fileDestinations.append(os.path.join(root,name))
                filnavn.append(os.path.splitext(name)[0])
                created.append(strftime("%d/%m/%Y", gmtime(os.path.getctime(os.path.join(root,name)))))
                
                # if not name.endswith(excludeExtensions):
                #     fileDestinations.append(os.path.join(root,name))
                #     filnavn.append(os.path.splitext(name)[0])
                #     created.append(strftime("%d/%m/%Y", gmtime(os.path.getctime(os.path.join(root,name)))))
# fileDestinations = list(dict.fromkeys(fileDestinations))
# filnavn = list(dict.fromkeys(filnavn))
# created = list(dict.fromkeys(created))


today=strftime("%d/%m/%Y")

# set col values
colDescription=1
colDato=2
colLink=3

# set row values
rowTitle=1
rowstart=2
bd = Side(style='thin', color="000000")

# Lock the all above and to the left of cell value
sheet.freeze_panes = sheet['A3']



# Title
sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)
sheet.cell(row=rowTitle, column=1).value = project+' - List of drawings - '+today
sheet.cell(row=rowTitle, column=1).font = Font(size='16', bold=True)
sheet.cell(row=rowTitle, column=1).alignment = Alignment(horizontal='center')

# Link
sheet.cell(row=rowstart, column=colLink).value = 'Link'
sheet.cell(row=rowstart, column=colLink).font = Font(bold=True)
sheet.cell(row=rowstart, column=colLink).border = Border(bottom=bd, top=bd)
sheet.column_dimensions[get_column_letter(colLink)].width = '6.8'

# Date
sheet.cell(row=rowstart, column=colDato).value = 'Date'
sheet.cell(row=rowstart, column=colDato).font = Font(bold=True)
sheet.cell(row=rowstart, column=colDato).border = Border(bottom=bd, top=bd)
sheet.column_dimensions[get_column_letter(colDato)].width = '10.7'

#Description
sheet.cell(row=rowstart, column=colDescription).value ='Description'
sheet.cell(row=rowstart, column=colDescription).font = Font(bold=True)
sheet.cell(row=rowstart, column=colDescription).border = Border(bottom=bd, top=bd)
sheet.column_dimensions[get_column_letter(colDescription)].width = '71.5'


i=rowstart+1
for File in fileDestinations:
    sheet.cell(row=i, column=colLink).value = ('=HYPERLINK("'+File+'", "PDF")')
    sheet.cell(row=i, column=colLink).style = 'Hyperlink'               
    i+=1

i=rowstart+1
for des in filnavn:
    sheet.cell(row=i, column=colDescription).value = (des)
    i+=1

i=rowstart+1
for dato in created:
    sheet.cell(row=i, column=colDato).value = (dato)
    sheet.cell(row=i, column=colDato).number_format = 'dd-mm-yyyy'
    i+=1
#timestr = time.strftime("%Y%m%d-%H%M")

# get max row count
max_row=sheet.max_row


sheet.auto_filter.ref= 'A2:C'+str(max_row)

##Autofit column
# column_widths = []
# for row in sheet.iter_rows():
#     for i, cell in enumerate(row):
#         try:
#             column_widths[i] = max(column_widths[i], len(str(cell.value)))
#         except IndexError:
#             column_widths.append(len(str(cell.value)))

# for i, column_width in enumerate(column_widths):
#     sheet.column_dimensions[get_column_letter(i + 1)].width = column_width



#save excel   
sheet.title='List of drawings'    
saveName=destPath+'\\'+project+' - List of drawings.xlsx'

yesnosave=tk.messagebox.askyesno(title='List of Drawings', message='Do you want to save "'+project+'.xlsx"?')

if yesnosave == 1:
    wb.save(saveName)
else:
    os._exit(1) 

yesno=tk.messagebox.askyesno(title='List of Drawings', message='Do you want to open the file "'+project+'.xlsx" in Excel?')

if yesno == 1:
    #xl=win32com.client.Dispatch("Excel.Application")
    xl=win32.Dispatch("Excel.Application")
    xl.Visible = True
    xl.Workbooks.Open(saveName)
else:
    tk.messagebox.showinfo(title='Lits of drwaings', message='Done')


#Pyinstaller --onefile ListOfDrawings.py